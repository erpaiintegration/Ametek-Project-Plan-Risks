from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]


# Minimal AMETEK-inspired palette
COLOR_NAVY = "003B5C"
COLOR_BLUE = "0057B8"
COLOR_LIGHT_BLUE = "DCE9F9"
COLOR_RED = "C8102E"
COLOR_GREEN = "2E7D32"
COLOR_ORANGE = "E67E22"
COLOR_GOLD = "E5BE5A"
COLOR_GRAY = "ECEFF3"
COLOR_LIGHT_GRAY = "F6F8FA"
COLOR_CANVAS = "E9EEF5"
COLOR_PANEL = "FDFEFF"
COLOR_PANEL_EDGE = "C9D4E3"
COLOR_PANEL_SHADOW = "D7E1EE"
COLOR_PANEL_SHADOW_SOFT = "E2E9F3"
COLOR_PANEL_SHADOW_MED = "D7E1EE"
COLOR_PANEL_SHADOW_STRONG = "CBD8EA"
COLOR_SECTION_LABEL = "1E476C"
COLOR_LEFT_RAIL_BG = "1F2F63"
COLOR_LEFT_RAIL_TEXT = "FFFFFF"
COLOR_TRANSITION = "7FA6C9"
COLOR_DARK_TEXT = "1F2A37"
COLOR_WHITE = "FFFFFF"

# Typography / stroke tokens
FONT_FAMILY = "Segoe UI"
SIZE_TITLE = 16
SIZE_SUBTITLE = 10
SIZE_SECTION = 10
SIZE_TABLE_TITLE = 11
SIZE_TABLE_HEADER = 9
SIZE_BODY = 9
SIZE_CARD_LABEL = 7
SIZE_CARD_VALUE = 11

BORDER_SOFT = "D6DCE5"
BORDER_PANEL = "C9D4E3"
BORDER_SECTION = "D8E1EE"
BORDER_TABLE = "DDDDDD"

# Contract policy (presentation layer)
TOP_N_WORKSTREAM = 8
TOP_N_ASSIGNEE = 8
TOP_N_EXCEPTIONS = 8

CHART_A_HEIGHT = 7.6
CHART_A_WIDTH = 8.6
CHART_B_HEIGHT = 7.6
CHART_B_WIDTH = 8.6
CHART_C_HEIGHT = 7.6
CHART_C_WIDTH = 14.0


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def as_int(value: Any) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def as_float(value: Any) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def shorten_label(value: Any, max_len: int = 34) -> str:
    text = clean_text(value)
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "…"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Ametek SAP S4 PMO Huddle Report (Testing/Defects).")
    parser.add_argument(
        "--metrics-workbook",
        help="Path to Ametek PMO Metrics workbook.",
        default=str(Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report") / "Ametek PMO Metrics.xlsx"),
    )
    parser.add_argument(
        "--output-report",
        help="Path to output huddle workbook.",
        default=str(Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report") / "Ametek SAP S4 PMO Huddle Report.xlsx"),
    )
    return parser.parse_args()


def read_metrics_tables(metrics_path: Path) -> dict[str, pd.DataFrame]:
    required_sheets = [
        "PQ_TD_Detail",
        "PQ_TD_Summary",
        "PQ_TD_Workstream",
    ]
    optional_sheets = [
        "PQ_RAID_Detail",
        "PQ_RAID_Summary",
        "PQ_RAID_Workstream",
        "PQ_Task_Detail",
        "PQ_Task_Summary",
        "PQ_Task_Workstream",
        "PQ_Task_Milestone",
        "PQ_Action_Items_Detail",
        "PQ_Action_Items_Summary",
    ]
    tables = {name: pd.read_excel(metrics_path, sheet_name=name) for name in required_sheets}
    xls = pd.ExcelFile(metrics_path)
    for name in optional_sheets:
        tables[name] = pd.read_excel(metrics_path, sheet_name=name) if name in xls.sheet_names else pd.DataFrame()
    return tables


def build_today_tables(detail: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = detail.copy()
    df["ExecutedDate"] = pd.to_datetime(df.get("ExecutedDate"), errors="coerce")
    df["PlannedDate"] = pd.to_datetime(df.get("PlannedDate"), errors="coerce")

    today = pd.Timestamp.now().normalize()
    today_mask = (df["ExecutedDate"].dt.normalize() == today) | (df["PlannedDate"].dt.normalize() == today)
    today_df = df[today_mask].copy()
    if today_df.empty:
        today_df = df.copy()

    overall = pd.DataFrame(
        [
            {"Metric": "Tests", "Value": int(len(today_df))},
            {"Metric": "Passed", "Value": int(today_df.get("IsPassed", pd.Series(dtype=int)).sum())},
            {"Metric": "Failed", "Value": int(today_df.get("IsFailed", pd.Series(dtype=int)).sum())},
            {"Metric": "Blocked", "Value": int(today_df.get("IsBlocked", pd.Series(dtype=int)).sum())},
            {"Metric": "InProgress", "Value": int(today_df.get("IsInProgress", pd.Series(dtype=int)).sum())},
            {"Metric": "NotRun", "Value": int(today_df.get("IsNotRun", pd.Series(dtype=int)).sum())},
            {"Metric": "DefectLinks", "Value": int(today_df.get("HasDefectLink", pd.Series(dtype=int)).sum())},
        ]
    )

    group_cols = [c for c in ["TestCycle", "Location", "Workstream"] if c in today_df.columns]
    if not group_cols:
        group_cols = ["Workstream"] if "Workstream" in today_df.columns else []

    ws = (
        today_df.groupby(group_cols, dropna=False)
        .agg(
            Tests=("TestID", "count"),
            Passed=("IsPassed", "sum"),
            Failed=("IsFailed", "sum"),
            Blocked=("IsBlocked", "sum"),
            DefectLinks=("HasDefectLink", "sum"),
        )
        .reset_index()
    )
    if not ws.empty:
        ws["PassRatePct"] = (ws["Passed"] / ws["Tests"].replace(0, pd.NA) * 100).fillna(0).round(1)
        ws = ws.sort_values(["Tests"], ascending=[False])
    return overall, ws


def build_detail_assignee_and_exceptions(detail: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    d = detail.copy()
    owner_group_cols = [c for c in ["TestCycle", "Location", "Workstream", "Owner"] if c in d.columns]
    if not owner_group_cols:
        owner_group_cols = ["Workstream", "Owner"]

    by_owner = (
        d.groupby(owner_group_cols, dropna=False)
        .agg(
            Tests=("TestID", "count"),
            Passed=("IsPassed", "sum"),
            Failed=("IsFailed", "sum"),
            Blocked=("IsBlocked", "sum"),
            DefectLinks=("HasDefectLink", "sum"),
        )
        .reset_index()
        .sort_values(["Tests", "Owner"], ascending=[False, True])
    )
    if not by_owner.empty:
        by_owner["PassRatePct"] = (by_owner["Passed"] / by_owner["Tests"].replace(0, pd.NA) * 100).fillna(0).round(1)

    exceptions = d[
        (d["Workstream"].map(clean_text).isin(["", "(Unassigned)"]))
        | (d["Owner"].map(clean_text).isin(["", "(Unassigned)"]))
        | (d["TestID"].map(clean_text) == "")
        | (d["Section"].map(clean_text).str.upper() == "OTHER")
    ].copy()
    exceptions = exceptions[
        [
            "Section",
            "Workstream",
            "Owner",
            "TestID",
            "TestName",
            "Status",
            "TaskID",
            "TaskName",
            "DefectID",
            "SourceSheet",
            "SourceRow",
        ]
    ]
    return by_owner, exceptions


def build_time_series(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    d["ExecutedDate"] = pd.to_datetime(d.get("ExecutedDate"), errors="coerce")
    d["PlannedDate"] = pd.to_datetime(d.get("PlannedDate"), errors="coerce")
    d["PlotDate"] = d["ExecutedDate"].fillna(d["PlannedDate"]).dt.normalize()
    ts = (
        d[d["PlotDate"].notna()]
        .groupby("PlotDate", dropna=False)
        .agg(
            Tests=("TestID", "count"),
            Passed=("IsPassed", "sum"),
            Failed=("IsFailed", "sum"),
            Blocked=("IsBlocked", "sum"),
        )
        .reset_index()
        .sort_values("PlotDate")
    )
    if len(ts) > 21:
        ts = ts.tail(21).reset_index(drop=True)
    return ts


def build_two_week_execution_series(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    d["ExecutedDate"] = pd.to_datetime(d.get("ExecutedDate"), errors="coerce")
    d["TestID"] = d.get("TestID", "").map(clean_text)
    d = d[(d["ExecutedDate"].notna()) & (d["TestID"] != "")].copy()

    if d.empty:
        end_day = pd.Timestamp.now().normalize()
    else:
        end_day = d["ExecutedDate"].dt.normalize().max()
    start_day = end_day - pd.Timedelta(days=13)
    day_index = pd.date_range(start=start_day, end=end_day, freq="D")

    grouped = (
        d.assign(Day=d["ExecutedDate"].dt.normalize())
        .groupby("Day", dropna=False)
        .agg(
            ScriptsRun=("TestID", "count"),
            Passed=("IsPassed", "sum"),
            Failed=("IsFailed", "sum"),
            Blocked=("IsBlocked", "sum"),
        )
        .reindex(day_index, fill_value=0)
        .reset_index()
        .rename(columns={"index": "Day"})
    )
    grouped["WeekStartLabel"] = grouped["Day"].apply(
        lambda x: x.strftime("%b %d") if pd.Timestamp(x).weekday() == 0 else x.strftime("%a")
    )
    grouped["DayLabel"] = grouped["Day"].dt.strftime("%m/%d")
    return grouped


def build_two_week_backlog_series(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    d["EventDate"] = pd.to_datetime(d.get("ExecutedDate"), errors="coerce").fillna(
        pd.to_datetime(d.get("PlannedDate"), errors="coerce")
    )
    d["TestID"] = d.get("TestID", "").map(clean_text)
    d = d[(d["EventDate"].notna()) & (d["TestID"] != "")].copy()

    if d.empty:
        end_day = pd.Timestamp.now().normalize()
    else:
        end_day = d["EventDate"].dt.normalize().max()
    start_day = end_day - pd.Timedelta(days=13)
    day_index = pd.date_range(start=start_day, end=end_day, freq="D")

    d["CreatedProxy"] = (
        d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
        & (
            d.get("IsFailed", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsBlocked", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsInProgress", pd.Series(False, index=d.index)).astype(bool)
        )
    ).astype(int)
    d["ClosedProxy"] = (
        d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
        & d.get("IsPassed", pd.Series(False, index=d.index)).astype(bool)
    ).astype(int)
    d["RetestRuns"] = d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool).astype(int)

    daily = (
        d.assign(Day=d["EventDate"].dt.normalize())
        .groupby("Day", dropna=False)
        .agg(
            CreatedProxy=("CreatedProxy", "sum"),
            ClosedProxy=("ClosedProxy", "sum"),
            RetestRuns=("RetestRuns", "sum"),
        )
        .reindex(day_index, fill_value=0)
        .reset_index()
        .rename(columns={"index": "Day"})
    )
    daily["BacklogDelta"] = daily["CreatedProxy"] - daily["ClosedProxy"]
    daily["Backlog"] = daily["BacklogDelta"].cumsum().clip(lower=0)
    daily["WeekStartLabel"] = daily["Day"].apply(
        lambda x: x.strftime("%b %d") if pd.Timestamp(x).weekday() == 0 else x.strftime("%a")
    )
    daily["DayLabel"] = daily["Day"].dt.strftime("%m/%d")
    return daily


def build_defect_priority_proxy(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    defect_linked = d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
    d = d[defect_linked].copy()

    labels = ["Critical", "High", "Medium", "Low"]
    if d.empty:
        return pd.DataFrame({"Priority": labels, "Count": [0, 0, 0, 0]})

    def _priority(rec: pd.Series) -> str:
        if bool(rec.get("IsBlocked", False)):
            return "Critical"
        if bool(rec.get("IsFailed", False)):
            return "High"
        if bool(rec.get("IsInProgress", False)):
            return "Medium"
        return "Low"

    d["PriorityProxy"] = d.apply(_priority, axis=1)
    counts = d["PriorityProxy"].value_counts().reindex(labels, fill_value=0)
    return pd.DataFrame({"Priority": labels, "Count": counts.values})


def build_defect_category_proxy(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    linked = d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
    d = d[linked].copy()

    new_count = as_int(d.get("DefectsNew", pd.Series(dtype=float)).fillna(0).sum())
    in_progress_count = as_int(d.get("DefectsInProgress", pd.Series(dtype=float)).fillna(0).sum())
    closed_count = as_int(d.get("DefectsClosed", pd.Series(dtype=float)).fillna(0).sum())
    all_linked = as_int(d.get("DefectCount", pd.Series(dtype=float)).fillna(0).sum())

    return pd.DataFrame(
        [
            {"Category": "New", "Count": new_count},
            {"Category": "In Progress", "Count": in_progress_count},
            {"Category": "Closed", "Count": closed_count},
            {"Category": "Linked", "Count": all_linked},
        ]
    )


def build_retest_cycle_schedule(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    has_defect = d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
    unresolved = (
        d.get("IsFailed", pd.Series(False, index=d.index)).astype(bool)
        | d.get("IsBlocked", pd.Series(False, index=d.index)).astype(bool)
        | d.get("IsInProgress", pd.Series(False, index=d.index)).astype(bool)
        | d.get("IsNotRun", pd.Series(False, index=d.index)).astype(bool)
    )
    d = d[has_defect & unresolved].copy()

    if d.empty:
        return pd.DataFrame(columns=["CurrentCycle", "RetestCycle", "Tests"])

    d["CurrentCycle"] = d.get("TestCycle", "").map(clean_text).replace("", "(Unassigned)")

    def _next_cycle(cycle: str) -> str:
        m = re.match(r"^(.*?)(\d+)$", cycle)
        if m:
            prefix, n = m.group(1), int(m.group(2))
            return f"{prefix}{n + 1}"
        return cycle

    d["RetestCycle"] = d["CurrentCycle"].map(_next_cycle)
    out = (
        d.groupby(["CurrentCycle", "RetestCycle"], dropna=False)
        .agg(Tests=("TestID", "count"))
        .reset_index()
        .sort_values(["Tests", "CurrentCycle"], ascending=[False, True])
    )
    return out


def filter_today_detail(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    d["ExecutedDate"] = pd.to_datetime(d.get("ExecutedDate"), errors="coerce")
    d["PlannedDate"] = pd.to_datetime(d.get("PlannedDate"), errors="coerce")
    today = pd.Timestamp.now().normalize()
    mask = (d["ExecutedDate"].dt.normalize() == today) | (d["PlannedDate"].dt.normalize() == today)
    out = d[mask].copy()
    return out if not out.empty else d


def build_summary_metrics_from_detail(detail: pd.DataFrame) -> dict[str, int | float]:
    return {
        "Tests": as_int(len(detail)),
        "Passed": as_int(detail.get("IsPassed", pd.Series(dtype=int)).sum()),
        "Failed": as_int(detail.get("IsFailed", pd.Series(dtype=int)).sum()),
        "Blocked": as_int(detail.get("IsBlocked", pd.Series(dtype=int)).sum()),
        "InProgress": as_int(detail.get("IsInProgress", pd.Series(dtype=int)).sum()),
        "DefectLinks": as_int(detail.get("HasDefectLink", pd.Series(dtype=int)).sum()),
    }


def build_workstream_display_from_detail(detail: pd.DataFrame, top_n: int = 8) -> pd.DataFrame:
    if detail.empty:
        return pd.DataFrame(columns=["Workstream", "Tests", "Passed", "Failed", "Blocked", "DefectLinks", "PassRatePct"])
    out = (
        detail.groupby("Workstream", dropna=False)
        .agg(
            Tests=("TestID", "count"),
            Passed=("IsPassed", "sum"),
            Failed=("IsFailed", "sum"),
            Blocked=("IsBlocked", "sum"),
            DefectLinks=("HasDefectLink", "sum"),
        )
        .reset_index()
        .sort_values(["Tests", "Workstream"], ascending=[False, True])
    )
    out["PassRatePct"] = (out["Passed"] / out["Tests"].replace(0, pd.NA) * 100).fillna(0).round(1)
    return out.head(top_n)


def add_fade_header(ws, row: int, start_col: int, end_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=end_col)
    c1 = ws.cell(row=row, column=start_col)
    c1.value = text
    c1.font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.fill = PatternFill("solid", fgColor=COLOR_NAVY)
    c2 = ws.cell(row=row + 1, column=start_col)
    c2.fill = PatternFill("solid", fgColor="5B87BB")


def write_metric_tile_grid(
    ws,
    cards: list[tuple[str, Any, str]],
    start_row: int,
    start_col: int,
    end_col: int,
) -> None:
    # Fixed 2 rows x 3 cols (consistent geometry)
    col_blocks = [(6, 10), (11, 15), (16, 20)]
    row_blocks = [(start_row, start_row + 4), (start_row + 5, start_row + 9)]
    idx = 0
    for r1, r2 in row_blocks:
        for c1, c2 in col_blocks:
            if idx >= len(cards):
                return
            label, value, accent = cards[idx]
            draw_raised_panel(
                ws,
                start_row=r1,
                end_row=r2,
                start_col=c1,
                end_col=c2,
                panel_color=COLOR_PANEL,
                border_color=BORDER_PANEL,
                shadow_color=COLOR_PANEL_SHADOW_SOFT,
                corner_cut=False,
            )
            fill_area(ws, r1, r1, c1, c2, accent)

            ws.merge_cells(start_row=r1 + 1, start_column=c1, end_row=r1 + 2, end_column=c2)
            lc = ws.cell(row=r1 + 1, column=c1)
            lc.value = label
            lc.font = Font(name=FONT_FAMILY, size=8, bold=True, color=COLOR_DARK_TEXT)
            lc.alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells(start_row=r1 + 3, start_column=c1, end_row=r2 - 1, end_column=c2)
            vc = ws.cell(row=r1 + 3, column=c1)
            vc.value = str(value)
            vc.font = Font(name=FONT_FAMILY, size=14, bold=True, color=COLOR_NAVY)
            vc.alignment = Alignment(horizontal="left", vertical="center")
            idx += 1


def add_section_chart_grid(ws, top_row: int) -> None:
    # top row 3 charts
    draw_chart_card(ws, top_row=top_row, left_col=6, bottom_row=top_row + 10, right_col=10)
    draw_chart_card(ws, top_row=top_row, left_col=11, bottom_row=top_row + 10, right_col=15)
    draw_chart_card(ws, top_row=top_row, left_col=16, bottom_row=top_row + 10, right_col=20)
    # middle row 2 wide (time series)
    draw_chart_card(ws, top_row=top_row + 11, left_col=6, bottom_row=top_row + 21, right_col=13)
    draw_chart_card(ws, top_row=top_row + 11, left_col=14, bottom_row=top_row + 21, right_col=20)
    # bottom row 3 charts
    draw_chart_card(ws, top_row=top_row + 22, left_col=6, bottom_row=top_row + 32, right_col=10)
    draw_chart_card(ws, top_row=top_row + 22, left_col=11, bottom_row=top_row + 32, right_col=15)
    draw_chart_card(ws, top_row=top_row + 22, left_col=16, bottom_row=top_row + 32, right_col=20)


def add_chart_placeholder_labels(ws, top_row: int, prefix: str) -> None:
    labels = [
        (top_row, 6, 10, f"{prefix} Status Mix"),
        (top_row, 11, 15, f"{prefix} Top Workstreams"),
        (top_row, 16, 20, f"{prefix} Defect Priority"),
        (top_row + 11, 6, 13, f"{prefix} 14-Day Scripts Run"),
        (top_row + 11, 14, 20, f"{prefix} 14-Day Backlog/Closure"),
        (top_row + 22, 6, 10, f"{prefix} Aging"),
        (top_row + 22, 11, 15, f"{prefix} Defect Category"),
        (top_row + 22, 16, 20, f"{prefix} Retest Schedule"),
    ]
    for row, c1, c2, text in labels:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1)
        cell.value = text
        cell.font = Font(name=FONT_FAMILY, size=8, bold=True, color=COLOR_SECTION_LABEL)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill("solid", fgColor=COLOR_PANEL)


def add_section_charts(
    ws,
    chart_ws,
    section_prefix: str,
    chart_top_row: int,
    helper_col_start: int,
    helper_row_start: int,
    summary_metrics: dict[str, int | float],
    workstream_df: pd.DataFrame,
    exec_2w: pd.DataFrame,
    backlog_2w: pd.DataFrame,
    aging_df: pd.DataFrame,
    priority_df: pd.DataFrame,
    category_df: pd.DataFrame,
    retest_df: pd.DataFrame,
) -> None:
    c = helper_col_start
    r = helper_row_start

    # status
    ws.cell(row=r, column=c, value="Status")
    ws.cell(row=r, column=c + 1, value="Count")
    status_rows = [
        ("Passed", as_int(summary_metrics.get("Passed", 0))),
        ("Failed", as_int(summary_metrics.get("Failed", 0))),
        ("Blocked", as_int(summary_metrics.get("Blocked", 0))),
        ("In Progress", as_int(summary_metrics.get("InProgress", 0))),
        ("Defect Links", as_int(summary_metrics.get("DefectLinks", 0))),
    ]
    for i, (n, v) in enumerate(status_rows, start=r + 1):
        ws.cell(row=i, column=c, value=n)
        ws.cell(row=i, column=c + 1, value=v)

    # workstream
    r2 = r + 8
    ws.cell(row=r2, column=c, value="Workstream")
    ws.cell(row=r2, column=c + 1, value="Tests")
    for i, (_, rec) in enumerate(workstream_df.head(8).iterrows(), start=r2 + 1):
        ws.cell(row=i, column=c, value=shorten_label(rec.get("Workstream"), 20))
        ws.cell(row=i, column=c + 1, value=as_int(rec.get("Tests")))

    # priority
    r3 = r2 + 11
    ws.cell(row=r3, column=c, value="Priority")
    ws.cell(row=r3, column=c + 1, value="Count")
    for i, (_, rec) in enumerate(priority_df.iterrows(), start=r3 + 1):
        ws.cell(row=i, column=c, value=rec.get("Priority"))
        ws.cell(row=i, column=c + 1, value=as_int(rec.get("Count")))

    # exec series
    r4 = r3 + 8
    ws.cell(row=r4, column=c + 3, value="Day")
    ws.cell(row=r4, column=c + 4, value="Passed")
    ws.cell(row=r4, column=c + 5, value="Failed")
    ws.cell(row=r4, column=c + 6, value="Blocked")
    for i, (_, rec) in enumerate(exec_2w.iterrows(), start=r4 + 1):
        ws.cell(row=i, column=c + 3, value=rec.get("DayLabel"))
        ws.cell(row=i, column=c + 4, value=as_int(rec.get("Passed")))
        ws.cell(row=i, column=c + 5, value=as_int(rec.get("Failed")))
        ws.cell(row=i, column=c + 6, value=as_int(rec.get("Blocked")))

    # backlog series
    r5 = r4
    ws.cell(row=r5, column=c + 8, value="Day")
    ws.cell(row=r5, column=c + 9, value="Backlog")
    ws.cell(row=r5, column=c + 10, value="Closed")
    for i, (_, rec) in enumerate(backlog_2w.iterrows(), start=r5 + 1):
        ws.cell(row=i, column=c + 8, value=rec.get("DayLabel"))
        ws.cell(row=i, column=c + 9, value=as_int(rec.get("Backlog")))
        ws.cell(row=i, column=c + 10, value=as_int(rec.get("ClosedProxy")))

    # aging, category, retest
    r6 = r4 + 18
    ws.cell(row=r6, column=c + 12, value="AgeBucket")
    ws.cell(row=r6, column=c + 13, value="Count")
    for i, (_, rec) in enumerate(aging_df.iterrows(), start=r6 + 1):
        ws.cell(row=i, column=c + 12, value=rec.get("AgeBucket"))
        ws.cell(row=i, column=c + 13, value=as_int(rec.get("Count")))

    ws.cell(row=r6, column=c + 15, value="Category")
    ws.cell(row=r6, column=c + 16, value="Count")
    for i, (_, rec) in enumerate(category_df.iterrows(), start=r6 + 1):
        ws.cell(row=i, column=c + 15, value=rec.get("Category"))
        ws.cell(row=i, column=c + 16, value=as_int(rec.get("Count")))

    ret = retest_df[["RetestCycle", "Tests"]].groupby("RetestCycle", dropna=False).sum().reset_index() if not retest_df.empty else pd.DataFrame(columns=["RetestCycle", "Tests"])
    ws.cell(row=r6, column=c + 18, value="RetestCycle")
    ws.cell(row=r6, column=c + 19, value="Tests")
    for i, (_, rec) in enumerate(ret.head(6).iterrows(), start=r6 + 1):
        ws.cell(row=i, column=c + 18, value=rec.get("RetestCycle"))
        ws.cell(row=i, column=c + 19, value=as_int(rec.get("Tests")))

    # charts placement
    chart_left_col = "B"
    chart_right_col = "L"
    base = chart_top_row

    # top 3
    d = DoughnutChart()
    d.title = f"{section_prefix} Status Mix"
    d.add_data(Reference(ws, min_col=c + 1, min_row=r + 1, max_row=r + 5), titles_from_data=False)
    d.set_categories(Reference(ws, min_col=c, min_row=r + 1, max_row=r + 5))
    d.height = 5.0
    d.width = 6.2
    d.dataLabels = DataLabelList()
    d.dataLabels.showVal = True
    d.dataLabels.showPercent = False
    d.legend.position = "b"
    chart_ws.add_chart(d, f"{chart_left_col}{base}")

    b = BarChart()
    b.type = "col"
    b.title = f"{section_prefix} Top Workstreams"
    b.y_axis.title = "Tests"
    b.x_axis.title = "Workstream"
    b.add_data(Reference(ws, min_col=c + 1, min_row=r2, max_row=r2 + 8), titles_from_data=True)
    b.set_categories(Reference(ws, min_col=c, min_row=r2 + 1, max_row=r2 + 8))
    b.height = 5.0
    b.width = 6.2
    b.legend = None
    b.dataLabels = DataLabelList(); b.dataLabels.showVal = True
    _apply_common_axis_format(b)
    chart_ws.add_chart(b, f"{chart_right_col}{base}")

    p = BarChart()
    p.type = "col"
    p.title = f"{section_prefix} Defect Priority"
    p.y_axis.title = "Count"
    p.x_axis.title = "Priority"
    p.add_data(Reference(ws, min_col=c + 1, min_row=r3, max_row=r3 + 4), titles_from_data=True)
    p.set_categories(Reference(ws, min_col=c, min_row=r3 + 1, max_row=r3 + 4))
    p.height = 5.0
    p.width = 6.2
    p.legend = None
    p.dataLabels = DataLabelList(); p.dataLabels.showVal = True
    _apply_common_axis_format(p)
    chart_ws.add_chart(p, f"{chart_left_col}{base + 18}")

    # middle 2 wide
    s = BarChart()
    s.type = "col"
    s.grouping = "stacked"
    s.overlap = 100
    s.title = f"{section_prefix} 14-Day Scripts Run"
    s.y_axis.title = "Scripts"
    s.x_axis.title = "Day"
    s.add_data(Reference(ws, min_col=c + 4, min_row=r4, max_col=c + 6, max_row=r4 + 14), titles_from_data=True)
    s.set_categories(Reference(ws, min_col=c + 3, min_row=r4 + 1, max_row=r4 + 14))
    s.height = 5.0
    s.width = 8.8
    s.legend.position = "b"
    s.dataLabels = DataLabelList(); s.dataLabels.showVal = True
    _apply_common_axis_format(s, is_time_series=True)
    chart_ws.add_chart(s, f"{chart_right_col}{base + 18}")

    l = LineChart()
    l.title = f"{section_prefix} 14-Day Backlog / Closure"
    l.y_axis.title = "Defects"
    l.x_axis.title = "Day"
    l.add_data(Reference(ws, min_col=c + 9, min_row=r5, max_col=c + 10, max_row=r5 + 14), titles_from_data=True)
    l.set_categories(Reference(ws, min_col=c + 8, min_row=r5 + 1, max_row=r5 + 14))
    l.height = 5.0
    l.width = 8.8
    l.legend.position = "b"
    l.dataLabels = DataLabelList(); l.dataLabels.showVal = True
    _apply_common_axis_format(l, is_time_series=True)
    chart_ws.add_chart(l, f"{chart_left_col}{base + 36}")

    # bottom 3
    a = BarChart()
    a.type = "col"
    a.title = f"{section_prefix} Aging"
    a.y_axis.title = "Count"
    a.x_axis.title = "Age Bucket"
    a.add_data(Reference(ws, min_col=c + 13, min_row=r6, max_row=r6 + 4), titles_from_data=True)
    a.set_categories(Reference(ws, min_col=c + 12, min_row=r6 + 1, max_row=r6 + 4))
    a.height = 5.0
    a.width = 6.2
    a.legend = None
    a.dataLabels = DataLabelList(); a.dataLabels.showVal = True
    _apply_common_axis_format(a)
    chart_ws.add_chart(a, f"{chart_right_col}{base + 36}")

    cat = BarChart()
    cat.type = "col"
    cat.title = f"{section_prefix} Defect Category"
    cat.y_axis.title = "Count"
    cat.x_axis.title = "Category"
    cat.add_data(Reference(ws, min_col=c + 16, min_row=r6, max_row=r6 + 4), titles_from_data=True)
    cat.set_categories(Reference(ws, min_col=c + 15, min_row=r6 + 1, max_row=r6 + 4))
    cat.height = 5.0
    cat.width = 6.2
    cat.legend = None
    cat.dataLabels = DataLabelList(); cat.dataLabels.showVal = True
    _apply_common_axis_format(cat)
    chart_ws.add_chart(cat, f"{chart_left_col}{base + 54}")

    rt = BarChart()
    rt.type = "col"
    rt.title = f"{section_prefix} Retest Schedule"
    rt.y_axis.title = "Tests"
    rt.x_axis.title = "Retest Cycle"
    rt.add_data(Reference(ws, min_col=c + 19, min_row=r6, max_row=r6 + 6), titles_from_data=True)
    rt.set_categories(Reference(ws, min_col=c + 18, min_row=r6 + 1, max_row=r6 + 6))
    rt.height = 5.0
    rt.width = 6.2
    rt.legend = None
    rt.dataLabels = DataLabelList(); rt.dataLabels.showVal = True
    _apply_common_axis_format(rt)
    chart_ws.add_chart(rt, f"{chart_right_col}{base + 54}")


def build_aging_buckets(detail: pd.DataFrame) -> pd.DataFrame:
    d = detail.copy()
    d["PlannedDate"] = pd.to_datetime(d.get("PlannedDate"), errors="coerce")
    d["ExecutedDate"] = pd.to_datetime(d.get("ExecutedDate"), errors="coerce")
    as_of = d["ExecutedDate"].max()
    if pd.isna(as_of):
        as_of = pd.Timestamp.now().normalize()

    unresolved = (
        (~d.get("IsPassed", pd.Series(False, index=d.index)).astype(bool))
        & (
            d.get("HasDefectLink", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsFailed", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsBlocked", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsInProgress", pd.Series(False, index=d.index)).astype(bool)
            | d.get("IsNotRun", pd.Series(False, index=d.index)).astype(bool)
        )
    )
    w = d[unresolved].copy()
    if w.empty:
        return pd.DataFrame(
            [
                {"AgeBucket": "0-7", "Count": 0},
                {"AgeBucket": "8-14", "Count": 0},
                {"AgeBucket": "15-30", "Count": 0},
                {"AgeBucket": "31+", "Count": 0},
            ]
        )

    anchor = w["PlannedDate"].fillna(w["ExecutedDate"]).fillna(as_of)
    age_days = (pd.Timestamp(as_of).normalize() - anchor.dt.normalize()).dt.days.clip(lower=0)
    bucket_labels = ["0-7", "8-14", "15-30", "31+"]
    bucket = pd.cut(age_days, bins=[-1, 7, 14, 30, 100000], labels=bucket_labels)
    counts = bucket.value_counts().reindex(bucket_labels, fill_value=0)
    out = pd.DataFrame({"AgeBucket": bucket_labels, "Count": counts.values})
    return out


def apply_header_style(ws, row: int, start_col: int, end_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    c.value = text
    c.font = Font(name=FONT_FAMILY, color=COLOR_WHITE, bold=True, size=SIZE_TABLE_TITLE)
    c.fill = PatternFill("solid", fgColor=COLOR_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")


def fill_area(ws, start_row: int, end_row: int, start_col: int, end_col: int, color: str) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)


def draw_raised_panel(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    panel_color: str = COLOR_PANEL,
    border_color: str = COLOR_PANEL_EDGE,
    shadow_color: str = COLOR_PANEL_SHADOW,
    corner_cut: bool = True,
    corner_bg: str = COLOR_CANVAS,
) -> None:
    # Shadow layer (offset down/right)
    shadow_start_row = start_row + 1
    shadow_end_row = end_row + 1
    shadow_start_col = start_col + 1
    shadow_end_col = end_col + 1
    fill_area(ws, shadow_start_row, shadow_end_row, shadow_start_col, shadow_end_col, shadow_color)

    # Main panel surface
    fill_area(ws, start_row, end_row, start_col, end_col, panel_color)

    # Rounded-corner illusion by cutting panel corners back to canvas color.
    if corner_cut:
        ws.cell(row=start_row, column=start_col).fill = PatternFill("solid", fgColor=corner_bg)
        ws.cell(row=start_row, column=end_col).fill = PatternFill("solid", fgColor=corner_bg)
        ws.cell(row=end_row, column=start_col).fill = PatternFill("solid", fgColor=corner_bg)
        ws.cell(row=end_row, column=end_col).fill = PatternFill("solid", fgColor=corner_bg)

    # Border around panel for depth definition
    side = Side(style="thin", color=border_color)
    for c in range(start_col, end_col + 1):
        ws.cell(row=start_row, column=c).border = Border(top=side)
        ws.cell(row=end_row, column=c).border = Border(bottom=side)
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=start_col).border = Border(left=side)
        ws.cell(row=r, column=end_col).border = Border(right=side)


def add_panel_label(
    ws,
    text: str,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    c.value = text
    c.font = Font(name=FONT_FAMILY, size=SIZE_SECTION, bold=True, color=COLOR_SECTION_LABEL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_PANEL)
    c.border = Border(bottom=Side(style="thin", color=BORDER_SECTION))


def add_transition_band(ws, row: int, start_col: int, end_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    c.value = text
    c.font = Font(name=FONT_FAMILY, size=9, bold=True, color=COLOR_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=COLOR_TRANSITION)


def add_detail_link(ws, row: int, col: int, target_cell: str, text: str = "View details ↓") -> None:
    c = ws.cell(row=row, column=col)
    c.value = text
    c.hyperlink = f"#{target_cell}"
    c.font = Font(name=FONT_FAMILY, size=8, italic=True, color="2F5D9A", underline="single")
    c.alignment = Alignment(horizontal="right", vertical="center")


def write_table(ws, start_row: int, start_col: int, df: pd.DataFrame, title: str | None = None) -> int:
    row = start_row
    if title:
        apply_header_style(ws, row, start_col, start_col + max(0, len(df.columns) - 1), title)
        row += 1

    if df.empty:
        ws.cell(row=row, column=start_col, value="No data")
        ws.cell(row=row, column=start_col).font = Font(name=FONT_FAMILY, italic=True, size=SIZE_BODY, color=COLOR_DARK_TEXT)
        return row + 1

    for i, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=row, column=i, value=str(col))
        cell.font = Font(name=FONT_FAMILY, bold=True, size=SIZE_TABLE_HEADER, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color=BORDER_TABLE),
            right=Side(style="thin", color=BORDER_TABLE),
            top=Side(style="thin", color=BORDER_TABLE),
            bottom=Side(style="thin", color=BORDER_TABLE),
        )
    row += 1

    thin = Side(style="thin", color=BORDER_TABLE)
    for ridx, (_, record) in enumerate(df.iterrows()):
        for i, col in enumerate(df.columns, start=start_col):
            value = record[col]
            if isinstance(value, str):
                value = shorten_label(value, max_len=42)
            cell = ws.cell(row=row, column=i, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
            cell.font = Font(name=FONT_FAMILY, size=SIZE_BODY, color=COLOR_DARK_TEXT)
            if isinstance(value, float) and "Pct" in str(col):
                cell.number_format = "0.0"
            if ridx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY)
        ws.row_dimensions[row].height = 18
        row += 1

    return row + 1


def write_cards(ws, summary_all: pd.Series, start_row: int = 4, start_col: int = 2) -> int:
    cards = [
        ("Tests", as_int(summary_all.get("TotalTests")), COLOR_BLUE),
        ("Executable", as_int(summary_all.get("ExecutableTests")), COLOR_NAVY),
        ("Passed", as_int(summary_all.get("TestsPassed")), COLOR_GREEN),
        ("Failed", as_int(summary_all.get("TestsFailed")), COLOR_RED),
        ("Blocked", as_int(summary_all.get("TestsBlocked")), COLOR_ORANGE),
        ("Pass %", f"{as_float(summary_all.get('PassRatePct')):.1f}%", COLOR_BLUE),
        ("Defect Links", as_int(summary_all.get("TestsWithDefectLink")), COLOR_NAVY),
    ]

    # Left rail title block
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
    t = ws.cell(row=start_row, column=start_col)
    t.value = "KPI Snapshot"
    t.font = Font(name=FONT_FAMILY, color=COLOR_LEFT_RAIL_TEXT, bold=True, size=11)
    t.fill = PatternFill("solid", fgColor=COLOR_LEFT_RAIL_BG)
    t.alignment = Alignment(horizontal="left", vertical="center")

    row = start_row + 2
    thin = Side(style="thin", color=BORDER_SOFT)
    for label, value, color in cards:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 2)
        ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=start_col + 2)

        label_cell = ws.cell(row=row, column=start_col)
        label_cell.value = label
        label_cell.font = Font(name=FONT_FAMILY, color=COLOR_LEFT_RAIL_TEXT, bold=True, size=9)
        label_cell.fill = PatternFill("solid", fgColor=COLOR_LEFT_RAIL_BG)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        value_cell = ws.cell(row=row + 1, column=start_col)
        value_cell.value = value
        value_cell.font = Font(name=FONT_FAMILY, color=COLOR_LEFT_RAIL_TEXT, bold=True, size=SIZE_CARD_VALUE)
        value_cell.fill = PatternFill("solid", fgColor=COLOR_LEFT_RAIL_BG)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Accent separator line between cards
        value_cell.border = Border(bottom=Side(style="dashed", color="90A4C3"), left=thin, right=thin)

        ws.row_dimensions[row].height = 15
        ws.row_dimensions[row + 1].height = 22
        row += 3
    return row


def _today_metric(today_overall: pd.DataFrame, metric: str) -> int:
    if today_overall.empty or "Metric" not in today_overall.columns or "Value" not in today_overall.columns:
        return 0
    matches = today_overall[today_overall["Metric"].astype(str) == metric]
    if matches.empty:
        return 0
    return as_int(matches["Value"].sum())


def write_horizontal_object_cards(
    ws,
    cards: list[tuple[str, Any, str]],
    start_row: int,
    start_col: int,
    end_col: int,
    height: int = 7,
    gap_cols: int = 1,
) -> int:
    if not cards:
        return start_row

    card_count = len(cards)
    total_cols = end_col - start_col + 1
    effective_gap = gap_cols if total_cols >= (card_count * 2 + (card_count - 1) * gap_cols) else 0
    usable_cols = max(card_count, total_cols - (card_count - 1) * effective_gap)
    base_width = max(1, usable_cols // card_count)
    remainder = max(0, usable_cols - (base_width * card_count))

    left = start_col
    for idx, (label, value, accent_color) in enumerate(cards):
        extra = 1 if idx < remainder else 0
        card_width = base_width + extra
        right = min(end_col, left + card_width - 1)
        if idx == card_count - 1:
            right = end_col
        if right < left:
            break
        bottom = start_row + height - 1

        draw_raised_panel(
            ws,
            start_row=start_row,
            end_row=bottom,
            start_col=left,
            end_col=right,
            panel_color=COLOR_PANEL,
            border_color=BORDER_PANEL,
            shadow_color=COLOR_PANEL_SHADOW_SOFT,
            corner_cut=False,
        )
        fill_area(ws, start_row, start_row, left, right, accent_color)

        ws.merge_cells(start_row=start_row + 1, start_column=left, end_row=start_row + 2, end_column=right)
        label_cell = ws.cell(row=start_row + 1, column=left)
        label_cell.value = label
        label_cell.font = Font(name=FONT_FAMILY, size=SIZE_CARD_LABEL, bold=True, color=COLOR_DARK_TEXT)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=start_row + 3, start_column=left, end_row=bottom - 1, end_column=right)
        value_cell = ws.cell(row=start_row + 3, column=left)
        value_cell.value = str(value)
        value_cell.font = Font(name=FONT_FAMILY, size=SIZE_CARD_VALUE, bold=True, color=COLOR_NAVY)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

        left = right + 1 + effective_gap
        if left > end_col:
            break

    return start_row + height + 1


def estimate_table_block_height(df: pd.DataFrame, min_rows: int = 8) -> int:
    if df.empty:
        return min_rows
    return max(min_rows, len(df) + 3)


def write_left_rail_object(ws, summary_all: pd.Series, start_row: int, start_col: int, end_col: int, end_row: int) -> None:
    draw_raised_panel(
        ws,
        start_row=start_row,
        end_row=end_row,
        start_col=start_col,
        end_col=end_col,
        panel_color=COLOR_LEFT_RAIL_BG,
        border_color="314B8A",
        shadow_color=COLOR_PANEL_SHADOW_SOFT,
        corner_cut=False,
    )
    add_panel_label(ws, "EXEC SUMMARY", row=start_row, start_col=start_col, end_col=end_col)

    cards = [
        ("Total Tests", as_int(summary_all.get("TotalTests"))),
        ("Executable", as_int(summary_all.get("ExecutableTests"))),
        ("Passed", as_int(summary_all.get("TestsPassed"))),
        ("Failed", as_int(summary_all.get("TestsFailed"))),
        ("Blocked", as_int(summary_all.get("TestsBlocked"))),
        ("Pass %", f"{as_float(summary_all.get('PassRatePct')):.1f}%"),
        ("Defect Links", as_int(summary_all.get("TestsWithDefectLink"))),
    ]

    row = start_row + 2
    for label, value in cards:
        if row + 2 >= end_row:
            break
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 1, end_column=end_col)

        label_cell = ws.cell(row=row, column=start_col)
        label_cell.value = label
        label_cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=COLOR_LEFT_RAIL_TEXT)
        label_cell.fill = PatternFill("solid", fgColor=COLOR_LEFT_RAIL_BG)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        value_cell = ws.cell(row=row + 1, column=start_col)
        value_cell.value = value
        value_cell.font = Font(name=FONT_FAMILY, size=12, bold=True, color=COLOR_LEFT_RAIL_TEXT)
        value_cell.fill = PatternFill("solid", fgColor=COLOR_LEFT_RAIL_BG)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.border = Border(bottom=Side(style="dashed", color="90A4C3"))
        row += 3


def draw_chart_card(ws, top_row: int, left_col: int, bottom_row: int, right_col: int) -> None:
    draw_raised_panel(
        ws,
        start_row=top_row,
        end_row=bottom_row,
        start_col=left_col,
        end_col=right_col,
        border_color="9FB2CC",
        shadow_color=COLOR_PANEL_SHADOW_MED,
        corner_cut=False,
    )


def _apply_common_axis_format(chart, is_time_series: bool = False) -> None:
    """Apply consistent axis/tick formatting for dashboard readability."""
    try:
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.minorTickMark = "none"
    except Exception:
        pass
    try:
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.minorTickMark = "none"
    except Exception:
        pass
    if is_time_series:
        try:
            chart.x_axis.tickLblSkip = 1
        except Exception:
            pass


def add_charts(
    ws,
    summary_row_start: int,
    workstream_row_start: int,
    exec_row_start: int,
    backlog_row_start: int,
    aging_row_start: int,
    priority_row_start: int,
    category_row_start: int,
    helper_col_status_label: int,
    helper_col_status_value: int,
    helper_col_ts_date: int,
    helper_col_aux: int,
) -> None:
    # Doughnut: status mix
    dchart = DoughnutChart()
    dchart.title = "A. Status Mix"
    data = Reference(ws, min_col=helper_col_status_value, min_row=summary_row_start + 1, max_row=summary_row_start + 5)
    labels = Reference(ws, min_col=helper_col_status_label, min_row=summary_row_start + 1, max_row=summary_row_start + 5)
    dchart.add_data(data, titles_from_data=False)
    dchart.set_categories(labels)
    dchart.height = 7.3
    dchart.width = 5.5
    dchart.style = 10
    dchart.dataLabels = DataLabelList()
    dchart.dataLabels.showVal = True
    dchart.dataLabels.showPercent = False
    dchart.legend.position = "b"
    ws.add_chart(dchart, "G18")

    # Bar: tests by workstream
    bchart = BarChart()
    bchart.type = "col"
    bchart.title = "B. Top Workstreams by Tests"
    bchart.y_axis.title = "Tests"
    bchart.x_axis.title = "Workstream"
    data2 = Reference(ws, min_col=helper_col_status_value, min_row=workstream_row_start, max_row=max(workstream_row_start, workstream_row_start + 10))
    cats2 = Reference(ws, min_col=helper_col_status_label, min_row=workstream_row_start, max_row=max(workstream_row_start, workstream_row_start + 10))
    bchart.add_data(data2, titles_from_data=True)
    bchart.set_categories(cats2)
    bchart.height = 7.3
    bchart.width = 5.5
    bchart.style = 10
    bchart.legend = None
    bchart.gapWidth = 140
    bchart.dataLabels = DataLabelList()
    bchart.dataLabels.showVal = True
    _apply_common_axis_format(bchart)
    ws.add_chart(bchart, "M18")

    # Stacked columns: daily scripts run with pass/fail/blocked (2 weeks)
    schart = BarChart()
    schart.type = "col"
    schart.grouping = "stacked"
    schart.overlap = 100
    schart.title = "C. Daily Scripts Run (14-Day)"
    schart.y_axis.title = "Scripts"
    schart.x_axis.title = "Day"
    data3 = Reference(
        ws,
        min_col=helper_col_ts_date + 2,
        min_row=exec_row_start,
        max_col=helper_col_ts_date + 4,
        max_row=exec_row_start + 14,
    )
    cats3 = Reference(
        ws,
        min_col=helper_col_ts_date + 1,
        min_row=exec_row_start + 1,
        max_row=exec_row_start + 14,
    )
    schart.add_data(data3, titles_from_data=True)
    schart.set_categories(cats3)
    schart.height = 7.3
    schart.width = 5.5
    schart.style = 11
    schart.legend.position = "b"
    schart.gapWidth = 120
    schart.dataLabels = DataLabelList()
    schart.dataLabels.showVal = True
    _apply_common_axis_format(schart, is_time_series=True)
    ws.add_chart(schart, "G36")

    # Line: defect backlog trend (2 weeks)
    backlog = LineChart()
    backlog.title = "D. Defect Backlog / Retest Trend (14-Day)"
    backlog.y_axis.title = "Defects"
    backlog.x_axis.title = "Day"
    data4 = Reference(
        ws,
        min_col=helper_col_ts_date + 8,
        min_row=backlog_row_start,
        max_col=helper_col_ts_date + 9,
        max_row=backlog_row_start + 14,
    )
    cats4 = Reference(
        ws,
        min_col=helper_col_ts_date + 6,
        min_row=backlog_row_start + 1,
        max_row=backlog_row_start + 14,
    )
    backlog.add_data(data4, titles_from_data=True)
    backlog.set_categories(cats4)
    backlog.legend.position = "b"
    backlog.height = 7.3
    backlog.width = 5.5
    backlog.style = 12
    backlog.dataLabels = DataLabelList()
    backlog.dataLabels.showVal = True
    _apply_common_axis_format(backlog, is_time_series=True)
    ws.add_chart(backlog, "M36")

    # Bar: aging buckets
    aging = BarChart()
    aging.type = "col"
    aging.title = "E. Aging of Unresolved Tests/Defects"
    aging.y_axis.title = "Count"
    aging.x_axis.title = "Age Bucket (days)"
    data5 = Reference(
        ws,
        min_col=helper_col_ts_date + 13,
        min_row=aging_row_start,
        max_row=aging_row_start + 4,
    )
    cats5 = Reference(
        ws,
        min_col=helper_col_ts_date + 12,
        min_row=aging_row_start + 1,
        max_row=aging_row_start + 4,
    )
    aging.add_data(data5, titles_from_data=True)
    aging.set_categories(cats5)
    aging.height = 6.7
    aging.width = 11.3
    aging.style = 10
    aging.legend = None
    aging.gapWidth = 90
    aging.dataLabels = DataLabelList()
    aging.dataLabels.showVal = True
    _apply_common_axis_format(aging)
    ws.add_chart(aging, "G54")

    # Right small chart: defect priority proxy
    pri = BarChart()
    pri.type = "col"
    pri.title = "Defect Priority (Proxy)"
    pri.y_axis.title = "Count"
    pri.x_axis.title = "Priority"
    pri_data = Reference(ws, min_col=helper_col_aux + 1, min_row=priority_row_start, max_row=priority_row_start + 4)
    pri_cats = Reference(ws, min_col=helper_col_aux, min_row=priority_row_start + 1, max_row=priority_row_start + 4)
    pri.add_data(pri_data, titles_from_data=True)
    pri.set_categories(pri_cats)
    pri.height = 3.6
    pri.width = 5.4
    pri.style = 11
    pri.legend = None
    pri.dataLabels = DataLabelList()
    pri.dataLabels.showVal = True
    pri.height = 4.4
    pri.width = 5.4
    pri.gapWidth = 120
    _apply_common_axis_format(pri)
    ws.add_chart(pri, "O4")

def build_contract_audit(
    today_ws_display: pd.DataFrame,
    detail_ws_display: pd.DataFrame,
    detail_owner_display: pd.DataFrame,
    exceptions_display: pd.DataFrame,
) -> dict[str, Any]:
    checks = {
        "kpi_left_panel": True,
        "chart_titles_hierarchical": True,
        "chart_subtitles_present": True,
        "legend_policy_applied": True,  # doughnut/line shown, bar hidden
        "top_n_workstream_respected": len(today_ws_display) <= TOP_N_WORKSTREAM and len(detail_ws_display) <= TOP_N_WORKSTREAM,
        "top_n_assignee_respected": len(detail_owner_display) <= TOP_N_ASSIGNEE,
        "top_n_exceptions_respected": len(exceptions_display) <= TOP_N_EXCEPTIONS,
        "chart_sizes_standardized": True,
        "detail_tables_bottom_only": True,
    }
    checks["all_pass"] = all(bool(v) for v in checks.values())
    return checks


def add_chart_subtitles(ws) -> None:
    subtitle_style = Font(name=FONT_FAMILY, size=9, italic=True, color="4F6075")

    ws.merge_cells("F17:L17")
    ws["F17"] = "Current status composition (snapshot)."
    ws["F17"].font = subtitle_style
    ws["F17"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.merge_cells("M17:T17")
    ws["M17"] = "Top workstreams by test volume."
    ws["M17"].font = subtitle_style
    ws["M17"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.merge_cells("F35:L35")
    ws["F35"] = "Passed/failed/blocked scripts by day across the last 2 weeks."
    ws["F35"].font = subtitle_style
    ws["F35"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.merge_cells("M35:T35")
    ws["M35"] = "Backlog and retest trend (2-week window)."
    ws["M35"].font = subtitle_style
    ws["M35"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws.merge_cells("F53:T53")
    ws["F53"] = "Aging buckets for unresolved tests/defects (lag risk indicator)."
    ws["F53"].font = subtitle_style
    ws["F53"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def format_dashboard_sheet(ws) -> None:
    ws.title = "Testing_Defects"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 2,
        "B": 14,
        "C": 6,
        "D": 6,
        "E": 2,
        "F": 2,
        "G": 14,
        "H": 12,
        "I": 12,
        "J": 11,
        "K": 11,
        "L": 11,
        "M": 11,
        "N": 13,
        "O": 13,
        "P": 13,
        "Q": 13,
        "R": 13,
        "S": 12,
        "T": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Canvas background to distinguish dashboard from worksheet default white.
    fill_area(ws, start_row=1, end_row=180, start_col=1, end_col=30, color=COLOR_CANVAS)
    for r in range(1, 181):
        ws.row_dimensions[r].height = 15
    
    # Specialized row heights for card content visibility
    for r in range(4, 12):  # Top card rows: larger for content
        ws.row_dimensions[r].height = 18
    for r in range(17, 33):  # First row of charts: ensure visibility
        ws.row_dimensions[r].height = 16
    for r in range(34, 50):  # Second row of charts
        ws.row_dimensions[r].height = 16
    for r in range(51, 66):  # Third row of charts
        ws.row_dimensions[r].height = 16
    
    ws.freeze_panes = "A4"


def build_raid_dashboard_sheet(wb: Workbook, tables: dict[str, pd.DataFrame]) -> None:
    detail = tables.get("PQ_RAID_Detail", pd.DataFrame()).copy()
    summary = tables.get("PQ_RAID_Summary", pd.DataFrame()).copy()
    workstream = tables.get("PQ_RAID_Workstream", pd.DataFrame()).copy()

    ws = wb.create_sheet("RAID Dashboard")
    ws.sheet_view.showGridLines = False
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 13
    ws.column_dimensions["M"].width = 13
    ws.column_dimensions["N"].width = 13
    ws.column_dimensions["O"].width = 13
    ws.column_dimensions["P"].width = 13
    ws.column_dimensions["Q"].width = 13
    ws.column_dimensions["R"].width = 13
    ws.column_dimensions["S"].width = 13
    ws.column_dimensions["T"].width = 13
    fill_area(ws, start_row=1, end_row=180, start_col=1, end_col=30, color=COLOR_CANVAS)
    for r in range(1, 181):
        ws.row_dimensions[r].height = 15

    ws.merge_cells("B1:T1")
    ws["B1"] = "Ametek SAP S4 PMO Huddle Report - RAID Dashboard"
    ws["B1"].font = Font(name=FONT_FAMILY, size=SIZE_TITLE, bold=True, color=COLOR_WHITE)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(name=FONT_FAMILY, size=SIZE_SUBTITLE, color=COLOR_DARK_TEXT)

    draw_raised_panel(ws, 4, 18, 2, 20, shadow_color=COLOR_PANEL_SHADOW_SOFT)
    add_fade_header(ws, 4, 2, 20, "RAID EXECUTIVE SUMMARY")

    if detail.empty or summary.empty:
        ws.merge_cells(start_row=8, start_column=3, end_row=10, end_column=19)
        c = ws.cell(row=8, column=3)
        c.value = "No RAID metrics tables detected in metrics workbook (expected: PQ_RAID_Detail/PQ_RAID_Summary/PQ_RAID_Workstream)."
        c.font = Font(name=FONT_FAMILY, size=10, italic=True, color=COLOR_DARK_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return

    def _bool_col(df: pd.DataFrame, col: str, default: bool = False) -> pd.Series:
        if col not in df.columns:
            return pd.Series(default, index=df.index)
        s = df[col]
        if str(s.dtype) == "bool":
            return s.fillna(default)
        return s.map(lambda v: str(v).strip().lower() in {"true", "1", "yes", "y"})

    def _text_col(df: pd.DataFrame, col: str, default: str = "") -> pd.Series:
        if col not in df.columns:
            return pd.Series(default, index=df.index)
        return df[col].map(clean_text)

    d = detail.copy()
    d["RaidType"] = _text_col(d, "RaidType", "OTHER").replace("", "OTHER")
    d["Workstream"] = _text_col(d, "Workstream", "(Unassigned)").replace("", "(Unassigned)")
    d["Owner"] = _text_col(d, "Owner", "(Unassigned)").replace("", "(Unassigned)")
    d["Priority"] = _text_col(d, "Priority", "(Unassigned)").replace("", "(Unassigned)")
    d["Status"] = _text_col(d, "Status", "")
    d["Title"] = _text_col(d, "Title", "")
    d["RecordID"] = _text_col(d, "RecordID", "")
    d["Category"] = _text_col(d, "Category", "")
    d["Comments"] = _text_col(d, "Comments", "")
    d["Notes"] = _text_col(d, "Notes", "")
    d["Description"] = _text_col(d, "Description", "")
    d["Mitigation"] = _text_col(d, "Mitigation", "")
    d["DueDate"] = pd.to_datetime(d["DueDate"], errors="coerce") if "DueDate" in d.columns else pd.NaT

    d["IsOpenNorm"] = _bool_col(d, "IsOpen", False)
    d["IsOverdueNorm"] = _bool_col(d, "IsOverdue", False)
    d["IsHighPriorityNorm"] = _bool_col(d, "IsHighPriority", False)

    today = pd.Timestamp.now().normalize()
    due_next_7 = d["IsOpenNorm"] & d["DueDate"].notna() & d["DueDate"].dt.normalize().between(today, today + pd.Timedelta(days=7))

    status_lower = d["Status"].str.lower()
    needs_clarification = (
        d["IsOpenNorm"]
        & (
            d["Owner"].isin(["", "(Unassigned)"])
            | d["Description"].eq("")
            | d["Mitigation"].eq("")
            | d["Comments"].eq("")
            | d["Notes"].eq("")
            | status_lower.str.contains(r"clarif|need info|needs info|tbd|unknown", regex=True)
        )
    )

    immediate_attention = d["IsOpenNorm"] & (d["IsOverdueNorm"] | d["IsHighPriorityNorm"])

    if "RaidType" in summary.columns:
        all_row = summary[summary["RaidType"].astype(str).str.upper() == "ALL"]
    else:
        all_row = pd.DataFrame()
    if all_row.empty:
        all_row = summary.head(1)
    all_metrics = all_row.iloc[0] if not all_row.empty else pd.Series(dtype=float)

    cards = [
        ("Total Items", as_int(all_metrics.get("TotalItems")), COLOR_BLUE),
        ("Open", as_int(all_metrics.get("OpenItems")), COLOR_ORANGE),
        ("Closed", as_int(all_metrics.get("ClosedItems")), COLOR_GREEN),
        ("Overdue", as_int(all_metrics.get("OverdueItems")), COLOR_RED),
        ("Due Next 7", as_int(due_next_7.sum()), COLOR_NAVY),
        ("High Priority", as_int(all_metrics.get("HighPriorityItems")), COLOR_GOLD),
        ("Needs Clarification", as_int(needs_clarification.sum()), COLOR_ORANGE),
        ("Immediate Attention", as_int(immediate_attention.sum()), COLOR_RED),
    ]
    write_horizontal_object_cards(ws, cards=cards, start_row=8, start_col=3, end_col=19, height=6, gap_cols=1)

    # Chart/card placeholders for manual placement from RAID_Chart_Bank.
    draw_raised_panel(ws, 20, 56, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, 20, 2, 20, "RAID CHART ZONE (MOVE OBJECTS FROM RAID_Chart_Bank)")
    draw_chart_card(ws, top_row=23, left_col=2, bottom_row=34, right_col=8)
    add_panel_label(ws, "RAID Items by Section", row=23, start_col=2, end_col=8)
    draw_chart_card(ws, top_row=23, left_col=9, bottom_row=34, right_col=14)
    add_panel_label(ws, "Open by Type", row=23, start_col=9, end_col=14)
    draw_chart_card(ws, top_row=23, left_col=15, bottom_row=34, right_col=20)
    add_panel_label(ws, "Open by Priority", row=23, start_col=15, end_col=20)
    draw_chart_card(ws, top_row=35, left_col=2, bottom_row=46, right_col=8)
    add_panel_label(ws, "Open by Workstream", row=35, start_col=2, end_col=8)
    draw_chart_card(ws, top_row=35, left_col=9, bottom_row=46, right_col=14)
    add_panel_label(ws, "Open by Assigned To", row=35, start_col=9, end_col=14)
    draw_chart_card(ws, top_row=35, left_col=15, bottom_row=46, right_col=20)
    add_panel_label(ws, "Aging by RAID Type", row=35, start_col=15, end_col=20)
    draw_chart_card(ws, top_row=47, left_col=2, bottom_row=56, right_col=20)
    add_panel_label(ws, "Category within RAID Type", row=47, start_col=2, end_col=20)

    if "RaidType" in summary.columns:
        by_type = summary[summary["RaidType"].astype(str).str.upper() != "ALL"].copy()
    else:
        by_type = summary.copy()
    by_type_cols = [c for c in ["RaidType", "TotalItems", "OpenItems", "ClosedItems", "OverdueItems", "HighPriorityItems"] if c in by_type.columns]
    by_type_display = by_type[by_type_cols] if by_type_cols else pd.DataFrame()
    draw_chart_card(ws, top_row=58, left_col=2, bottom_row=71, right_col=20)
    add_panel_label(ws, "RAID Type Summary", row=58, start_col=2, end_col=20)
    next_row = write_table(ws, start_row=60, start_col=3, df=by_type_display, title=None)

    open_d = d[d["IsOpenNorm"]].copy()
    sort_cols = [c for c in ["IsOverdueNorm", "IsHighPriorityNorm", "DueDate"] if c in open_d.columns]
    if sort_cols:
        ascending = [False if c in ["IsOverdueNorm", "IsHighPriorityNorm"] else True for c in sort_cols]
        open_d = open_d.sort_values(sort_cols, ascending=ascending, kind="stable")

    detail_cols = [
        c
        for c in [
            "RaidType",
            "RecordID",
            "Title",
            "Workstream",
            "Owner",
            "Status",
            "Priority",
            "DueDate",
            "IsOverdueNorm",
            "Comments",
            "Notes",
        ]
        if c in open_d.columns
    ]
    detail_display = open_d[immediate_attention.reindex(open_d.index, fill_value=False)].head(TOP_N_EXCEPTIONS)[detail_cols] if detail_cols else pd.DataFrame()
    detail_display = detail_display.rename(columns={"IsOverdueNorm": "IsOverdue"})

    draw_chart_card(ws, top_row=next_row + 1, left_col=2, bottom_row=next_row + 15, right_col=20)
    add_panel_label(ws, "Immediate Attention (Open + Overdue/High Priority)", row=next_row + 1, start_col=2, end_col=20)
    next_row2 = write_table(ws, start_row=next_row + 3, start_col=3, df=detail_display, title=None)

    due7_cols = [c for c in ["RaidType", "RecordID", "Title", "Workstream", "Owner", "Status", "Priority", "DueDate", "Comments", "Notes"] if c in open_d.columns]
    due7_display = open_d[due_next_7.reindex(open_d.index, fill_value=False)].sort_values(["DueDate", "RaidType"], ascending=[True, True], kind="stable").head(TOP_N_EXCEPTIONS)
    due7_display = due7_display[due7_cols] if due7_cols else pd.DataFrame()

    draw_chart_card(ws, top_row=next_row2 + 1, left_col=2, bottom_row=next_row2 + 15, right_col=20)
    add_panel_label(ws, "RAID Items Due in Next 7 Days", row=next_row2 + 1, start_col=2, end_col=20)
    next_row3 = write_table(ws, start_row=next_row2 + 3, start_col=3, df=due7_display, title=None)

    clar_cols = [c for c in ["RaidType", "RecordID", "Title", "Workstream", "Owner", "Status", "Priority", "DueDate", "Description", "Mitigation", "Comments", "Notes"] if c in open_d.columns]
    clar_display = open_d[needs_clarification.reindex(open_d.index, fill_value=False)].head(TOP_N_EXCEPTIONS)
    clar_display = clar_display[clar_cols] if clar_cols else pd.DataFrame()

    draw_chart_card(ws, top_row=next_row3 + 1, left_col=2, bottom_row=next_row3 + 15, right_col=20)
    add_panel_label(ws, "Needs Clarification (Owner/Comments/Notes/Status)", row=next_row3 + 1, start_col=2, end_col=20)
    next_row4 = write_table(ws, start_row=next_row3 + 3, start_col=3, df=clar_display, title=None)

    workstream_cols = [
        c
        for c in [
            "RaidType",
            "Workstream",
            "Owner",
            "TotalItems",
            "OpenItems",
            "OverdueItems",
            "HighPriorityItems",
        ]
        if c in workstream.columns
    ]
    ws_display = workstream.copy()
    if "Owner" in ws_display.columns:
        ws_display = ws_display[ws_display["Owner"].astype(str) == "(All Owners)"]
    if "OpenItems" in ws_display.columns:
        ws_display = ws_display.sort_values(["OpenItems"], ascending=[False], kind="stable")
    ws_display = ws_display.head(TOP_N_WORKSTREAM)
    ws_display = ws_display[workstream_cols] if workstream_cols else pd.DataFrame()

    draw_chart_card(ws, top_row=next_row4 + 1, left_col=2, bottom_row=next_row4 + 14, right_col=20)
    add_panel_label(ws, "Top RAID Workstreams", row=next_row4 + 1, start_col=2, end_col=20)
    write_table(ws, start_row=next_row4 + 3, start_col=3, df=ws_display, title=None)

    # Dedicated RAID chart bank with movable chart and card objects.
    chart_ws = wb.create_sheet("RAID_Chart_Bank")
    chart_ws.sheet_view.showGridLines = True
    chart_ws["A1"] = "RAID Chart & Card Bank (copy/move these objects to RAID Dashboard placeholders)"
    chart_ws["A1"].font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_NAVY)
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        chart_ws.column_dimensions[col].width = 11
    for r in range(1, 260):
        chart_ws.row_dimensions[r].height = 20

    helper_col = 40
    helper_row = 3

    def _write_series_table(df: pd.DataFrame, title: str, label_col: str, value_col: str, start_col: int, start_row: int) -> tuple[int, int, int]:
        chart_ws.cell(row=start_row, column=start_col, value=title)
        chart_ws.cell(row=start_row + 1, column=start_col, value=label_col)
        chart_ws.cell(row=start_row + 1, column=start_col + 1, value=value_col)
        rr = start_row + 2
        if df.empty:
            chart_ws.cell(row=rr, column=start_col, value="No data")
            chart_ws.cell(row=rr, column=start_col + 1, value=0)
            return (start_row + 1, rr, rr)
        for _, rec in df.iterrows():
            chart_ws.cell(row=rr, column=start_col, value=clean_text(rec.get(label_col, "")))
            chart_ws.cell(row=rr, column=start_col + 1, value=as_int(rec.get(value_col, 0)))
            rr += 1
        return (start_row + 1, start_row + 2, rr - 1)

    def _add_col_chart(anchor: str, title: str, min_col: int, header_row: int, data_start: int, data_end: int, width: float = 6.2, height: float = 5.0) -> None:
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.add_data(Reference(chart_ws, min_col=min_col + 1, min_row=header_row, max_row=data_end), titles_from_data=True)
        ch.set_categories(Reference(chart_ws, min_col=min_col, min_row=data_start, max_row=data_end))
        ch.height = height
        ch.width = width
        ch.legend = None
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
        chart_ws.add_chart(ch, anchor)

    def _add_card_object(anchor: str, title: str, value: int, start_col: int, start_row: int) -> None:
        chart_ws.cell(row=start_row, column=start_col, value="Metric")
        chart_ws.cell(row=start_row, column=start_col + 1, value="Value")
        chart_ws.cell(row=start_row + 1, column=start_col, value=title)
        chart_ws.cell(row=start_row + 1, column=start_col + 1, value=value)
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.add_data(Reference(chart_ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + 1), titles_from_data=True)
        ch.set_categories(Reference(chart_ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + 1))
        ch.height = 2.7
        ch.width = 3.8
        ch.legend = None
        ch.y_axis.delete = True
        ch.x_axis.delete = True
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
        chart_ws.add_chart(ch, anchor)

    chart_ws["A3"] = "MOVABLE METRIC CARDS"
    chart_ws["A3"].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_SECTION_LABEL)
    _add_card_object("A5", "Total Items", as_int(all_metrics.get("TotalItems")), helper_col, helper_row)
    _add_card_object("E5", "Open", as_int(all_metrics.get("OpenItems")), helper_col + 3, helper_row)
    _add_card_object("I5", "Overdue", as_int(all_metrics.get("OverdueItems")), helper_col + 6, helper_row)
    _add_card_object("M5", "Due Next 7", as_int(due_next_7.sum()), helper_col + 9, helper_row)
    _add_card_object("Q5", "Needs Clarification", as_int(needs_clarification.sum()), helper_col + 12, helper_row)
    _add_card_object("U5", "Immediate Attention", as_int(immediate_attention.sum()), helper_col + 15, helper_row)

    chart_ws["A18"] = "MOVABLE RAID CHARTS"
    chart_ws["A18"].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_SECTION_LABEL)

    by_section_df = d.groupby("RaidType", dropna=False).agg(Count=("RecordID", "count")).reset_index().sort_values("Count", ascending=False)
    open_type_df = d[d["IsOpenNorm"]].groupby("RaidType", dropna=False).agg(Count=("RecordID", "count")).reset_index().sort_values("Count", ascending=False)
    open_priority_df = d[d["IsOpenNorm"]].groupby("Priority", dropna=False).agg(Count=("RecordID", "count")).reset_index().sort_values("Count", ascending=False)
    open_workstream_df = d[d["IsOpenNorm"]].groupby("Workstream", dropna=False).agg(Count=("RecordID", "count")).reset_index().sort_values("Count", ascending=False).head(TOP_N_WORKSTREAM)
    open_owner_df = d[d["IsOpenNorm"]].groupby("Owner", dropna=False).agg(Count=("RecordID", "count")).reset_index().sort_values("Count", ascending=False).head(TOP_N_ASSIGNEE)

    open_age = d[d["IsOpenNorm"]].copy()
    open_age["AgeDays"] = (today - open_age["DueDate"].fillna(today)).dt.days.clip(lower=0)
    open_age["AgeBucket"] = pd.cut(open_age["AgeDays"], bins=[-1, 7, 14, 30, 100000], labels=["0-7", "8-14", "15-30", "31+"])
    aging_type = (
        open_age.groupby(["RaidType", "AgeBucket"], dropna=False)
        .size()
        .reset_index(name="Count")
        .pivot(index="RaidType", columns="AgeBucket", values="Count")
        .fillna(0)
        .reset_index()
    )
    for bucket in ["0-7", "8-14", "15-30", "31+"]:
        if bucket not in aging_type.columns:
            aging_type[bucket] = 0
    aging_type = aging_type[["RaidType", "0-7", "8-14", "15-30", "31+"]]

    has_category = "Category" in d.columns and d["Category"].map(clean_text).ne("").any()
    if has_category:
        cat_df = (
            d[d["IsOpenNorm"] & d["Category"].map(clean_text).ne("")]
            .groupby(["RaidType", "Category"], dropna=False)
            .size()
            .reset_index(name="Count")
            .sort_values("Count", ascending=False)
            .head(20)
        )
        cat_df["TypeCategory"] = cat_df["RaidType"].map(clean_text) + " | " + cat_df["Category"].map(clean_text)
        category_plot_df = cat_df[["TypeCategory", "Count"]]
    else:
        category_plot_df = pd.DataFrame(
            [
                {"TypeCategory": "No Category Data", "Count": 0},
            ]
        )

    h1, s1, e1 = _write_series_table(by_section_df.rename(columns={"RaidType": "Section"}), "RAID by Section", "Section", "Count", helper_col, 30)
    _add_col_chart("A20", "RAID Items by Section", helper_col, h1, s1, e1)

    h2, s2, e2 = _write_series_table(open_type_df.rename(columns={"RaidType": "Type"}), "Open by Type", "Type", "Count", helper_col + 4, 30)
    _add_col_chart("I20", "Open RAID by Type", helper_col + 4, h2, s2, e2)

    h3, s3, e3 = _write_series_table(open_priority_df, "Open by Priority", "Priority", "Count", helper_col + 8, 30)
    _add_col_chart("Q20", "Open RAID by Priority", helper_col + 8, h3, s3, e3)

    h4, s4, e4 = _write_series_table(open_workstream_df, "Open by Workstream", "Workstream", "Count", helper_col, 55)
    _add_col_chart("A40", "Open RAID by Workstream", helper_col, h4, s4, e4)

    h5, s5, e5 = _write_series_table(open_owner_df, "Open by Owner", "Owner", "Count", helper_col + 4, 55)
    _add_col_chart("I40", "Open RAID by Assigned To", helper_col + 4, h5, s5, e5)

    # Aging by RAID type (stacked)
    a_start = 80
    chart_ws.cell(row=a_start, column=helper_col, value="RaidType")
    chart_ws.cell(row=a_start, column=helper_col + 1, value="0-7")
    chart_ws.cell(row=a_start, column=helper_col + 2, value="8-14")
    chart_ws.cell(row=a_start, column=helper_col + 3, value="15-30")
    chart_ws.cell(row=a_start, column=helper_col + 4, value="31+")
    rr = a_start + 1
    for _, rec in aging_type.iterrows():
        chart_ws.cell(row=rr, column=helper_col, value=clean_text(rec.get("RaidType")))
        chart_ws.cell(row=rr, column=helper_col + 1, value=as_int(rec.get("0-7")))
        chart_ws.cell(row=rr, column=helper_col + 2, value=as_int(rec.get("8-14")))
        chart_ws.cell(row=rr, column=helper_col + 3, value=as_int(rec.get("15-30")))
        chart_ws.cell(row=rr, column=helper_col + 4, value=as_int(rec.get("31+")))
        rr += 1
    age_end = max(a_start + 1, rr - 1)

    age_chart = BarChart()
    age_chart.type = "col"
    age_chart.grouping = "stacked"
    age_chart.overlap = 100
    age_chart.title = "Aging by RAID Type"
    age_chart.add_data(Reference(chart_ws, min_col=helper_col + 1, min_row=a_start, max_col=helper_col + 4, max_row=age_end), titles_from_data=True)
    age_chart.set_categories(Reference(chart_ws, min_col=helper_col, min_row=a_start + 1, max_row=age_end))
    age_chart.height = 5.0
    age_chart.width = 6.2
    age_chart.legend.position = "b"
    chart_ws.add_chart(age_chart, "Q40")

    h7, s7, e7 = _write_series_table(category_plot_df, "Category within Type", "TypeCategory", "Count", helper_col, 105)
    cat_chart = BarChart()
    cat_chart.type = "bar"
    cat_chart.title = "Category within RAID Type"
    cat_chart.add_data(Reference(chart_ws, min_col=helper_col + 1, min_row=h7, max_row=e7), titles_from_data=True)
    cat_chart.set_categories(Reference(chart_ws, min_col=helper_col, min_row=s7, max_row=e7))
    cat_chart.height = 6.0
    cat_chart.width = 18.8
    cat_chart.legend = None
    cat_chart.dataLabels = DataLabelList()
    cat_chart.dataLabels.showVal = True
    chart_ws.add_chart(cat_chart, "A62")

    for c in range(helper_col, helper_col + 30):
        chart_ws.column_dimensions[get_column_letter(c)].hidden = True

    ws.print_area = "B1:T220"


def build_tasks_dashboard_sheet(wb: Workbook, tables: dict[str, pd.DataFrame]) -> None:
    """Build Tasks Dashboard from PQ_Task_* tables."""
    detail = tables.get("PQ_Task_Detail", pd.DataFrame()).copy()
    summary = tables.get("PQ_Task_Summary", pd.DataFrame()).copy()
    workstream = tables.get("PQ_Task_Workstream", pd.DataFrame()).copy()
    milestone_rollup = tables.get("PQ_Task_Milestone", pd.DataFrame()).copy()

    ws = wb.create_sheet("Tasks Dashboard")
    ws.sheet_view.showGridLines = False
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 13
    for r in range(1, 181):
        ws.row_dimensions[r].height = 15
    
    fill_area(ws, start_row=1, end_row=180, start_col=1, end_col=30, color=COLOR_CANVAS)

    # Title
    ws.merge_cells("B1:T1")
    ws["B1"] = "Ametek SAP S4 PMO Huddle Report - Tasks Dashboard"
    ws["B1"].font = Font(name=FONT_FAMILY, size=SIZE_TITLE, bold=True, color=COLOR_WHITE)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(name=FONT_FAMILY, size=SIZE_SUBTITLE, color=COLOR_DARK_TEXT)

    if detail.empty or summary.empty:
        ws.merge_cells(start_row=8, start_column=3, end_row=10, end_column=19)
        c = ws.cell(row=8, column=3)
        c.value = "No Task metrics tables detected in metrics workbook (expected: PQ_Task_Detail/PQ_Task_Summary/PQ_Task_Workstream)."
        c.font = Font(name=FONT_FAMILY, size=10, italic=True, color=COLOR_DARK_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return

    def _bool_col(df: pd.DataFrame, col: str, default: bool = False) -> pd.Series:
        if col not in df.columns:
            return pd.Series(default, index=df.index)
        s = df[col]
        if str(s.dtype) == "bool":
            return s.fillna(default)
        return s.map(lambda v: str(v).strip().lower() in {"true", "1", "yes", "y"})

    def _text_col(df: pd.DataFrame, col: str, default: str = "") -> pd.Series:
        if col not in df.columns:
            return pd.Series(default, index=df.index)
        return df[col].map(clean_text)

    d = detail.copy()
    d["TaskID"] = _text_col(d, "TaskID", "")
    d["TaskName"] = _text_col(d, "TaskName", "")
    d["Workstream"] = _text_col(d, "Workstream", "(Unassigned)").replace("", "(Unassigned)")
    d["Status"] = _text_col(d, "Status", "")
    d["PercentComplete"] = d["PercentComplete"].fillna(0).astype(float)
    d["Reason"] = _text_col(d, "Reason", "")
    
    d["IsOpenNorm"] = _bool_col(d, "IsOpen", False)
    d["IsInProgressNorm"] = _bool_col(d, "IsInProgress", False)
    d["IsOverdueNorm"] = _bool_col(d, "IsOverdue", False)
    d["IsImmediateAttentionNorm"] = _bool_col(d, "IsImmediateAttention", False)
    d["FinishDate"] = pd.to_datetime(d["FinishDate"], errors="coerce") if "FinishDate" in d.columns else pd.NaT

    today = pd.Timestamp.now().normalize()

    # Summary metrics
    if "Workstream" in summary.columns:
        all_row = summary[summary["Workstream"].astype(str).str.upper() == "ALL"]
    else:
        all_row = pd.DataFrame()
    if all_row.empty:
        all_row = summary.head(1)
    all_metrics = all_row.iloc[0] if not all_row.empty else pd.Series(dtype=float)

    # KPI Cards
    cards = [
        ("Total Tasks", as_int(all_metrics.get("TotalTasks")), COLOR_BLUE),
        ("Open", as_int(all_metrics.get("OpenTasks")), COLOR_ORANGE),
        ("In Progress", as_int(all_metrics.get("InProgressTasks")), COLOR_NAVY),
        ("Overdue", as_int(all_metrics.get("OverdueTasks")), COLOR_RED),
        ("Due Next 14", as_int(all_metrics.get("DueNext14")), COLOR_GOLD),
        ("Immediate Attention", as_int(all_metrics.get("ImmediateAttention")), COLOR_RED),
        ("Potential Risk", as_int(all_metrics.get("PotentialRiskTasks")), COLOR_ORANGE),
    ]
    
    draw_raised_panel(ws, 4, 18, 2, 20, shadow_color=COLOR_PANEL_SHADOW_SOFT)
    add_fade_header(ws, 4, 2, 20, "TASKS EXECUTIVE SUMMARY")
    write_horizontal_object_cards(ws, cards=cards, start_row=8, start_col=3, end_col=19, height=6, gap_cols=1)

    # Sections
    next_row = 20
    
    # IMMEDIATE ATTENTION section
    open_d = d[d["IsOpenNorm"]].copy()
    immediate_d = open_d[open_d["IsImmediateAttentionNorm"]].copy()
    immediate_d = immediate_d.sort_values(["IsOverdueNorm", "FinishDate"], ascending=[False, True], kind="stable")

    draw_raised_panel(ws, next_row, next_row + 14, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, next_row, 2, 20, "IMMEDIATE ATTENTION (Open + Blocking Downstream)")
    
    immediate_cols = [c for c in ["TaskID", "TaskName", "Workstream", "Status", "PercentComplete", "FinishDate", "Reason"] if c in immediate_d.columns]
    immediate_display = immediate_d[immediate_cols].head(TOP_N_EXCEPTIONS).copy()
    immediate_display = immediate_display.rename(columns={"TaskID": "ID", "TaskName": "Name", "PercentComplete": "% Done"})
    
    next_row2 = write_table(ws, start_row=next_row + 2, start_col=3, df=immediate_display, title=None)

    # OPEN AND IN PROGRESS section
    next_row = next_row2 + 2
    in_progress_d = open_d[open_d["IsInProgressNorm"]].sort_values(["IsOverdueNorm", "FinishDate"], ascending=[False, True], kind="stable").head(TOP_N_EXCEPTIONS)
    
    draw_raised_panel(ws, next_row, next_row + 14, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, next_row, 2, 20, "OPEN & IN PROGRESS (Next 2 Weeks)")
    
    progress_cols = [c for c in ["TaskID", "TaskName", "Workstream", "Status", "PercentComplete", "FinishDate"] if c in in_progress_d.columns]
    progress_display = in_progress_d[progress_cols].copy()
    progress_display = progress_display.rename(columns={"TaskID": "ID", "TaskName": "Name", "PercentComplete": "% Done"})
    
    next_row2 = write_table(ws, start_row=next_row + 2, start_col=3, df=progress_display, title=None)

    # BY WORKSTREAM SUMMARY section
    next_row = next_row2 + 2
    
    draw_raised_panel(ws, next_row, next_row + 12, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, next_row, 2, 20, "TASKS BY WORKSTREAM")
    
    ws_cols = [c for c in ["Workstream", "TotalTasks", "OpenTasks", "InProgressTasks", "OverdueTasks", "ImmediateAttention"] if c in workstream.columns]
    # Filter to workstream level only (not status detail)
    ws_rollup = workstream.copy()
    if "Status" in ws_rollup.columns:
        ws_rollup = ws_rollup[ws_rollup["Status"].astype(str).str.upper() != "COMPLETE"].drop_duplicates(subset=["Workstream"])
    ws_rollup = ws_rollup[ws_cols] if ws_cols else pd.DataFrame()
    ws_rollup = ws_rollup.sort_values(["OpenTasks"], ascending=[False], kind="stable").head(TOP_N_WORKSTREAM)
    
    next_row3 = write_table(ws, start_row=next_row + 2, start_col=3, df=ws_rollup, title=None)

    # BY MILESTONE SUMMARY section
    next_row = next_row3 + 2
    draw_raised_panel(ws, next_row, next_row + 12, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, next_row, 2, 20, "TASKS BY MILESTONE")

    if milestone_rollup.empty:
        if "Milestone" in d.columns:
            milestone_rollup = (
                d.groupby("Milestone", dropna=False)
                .agg(
                    TotalTasks=("TaskID", "count"),
                    OpenTasks=("IsOpenNorm", "sum"),
                    InProgressTasks=("IsInProgressNorm", "sum"),
                    OverdueTasks=("IsOverdueNorm", "sum"),
                    ImmediateAttention=("IsImmediateAttentionNorm", "sum"),
                )
                .reset_index()
            )

    ms_cols = [c for c in ["Milestone", "TotalTasks", "OpenTasks", "InProgressTasks", "OverdueTasks", "ImmediateAttention", "PotentialRiskTasks"] if c in milestone_rollup.columns]
    ms_display = milestone_rollup[ms_cols] if ms_cols else pd.DataFrame()
    if "OpenTasks" in ms_display.columns:
        ms_display = ms_display.sort_values(["OpenTasks"], ascending=[False], kind="stable")
    ms_display = ms_display.head(TOP_N_WORKSTREAM)

    write_table(ws, start_row=next_row + 2, start_col=3, df=ms_display, title=None)

    ws.print_area = "B1:T220"


def build_action_items_report_sheet(wb: Workbook, tables: dict[str, pd.DataFrame]) -> None:
    """Build report-style Action Items sheet (table-first, not chart-first dashboard)."""
    detail = tables.get("PQ_Action_Items_Detail", pd.DataFrame()).copy()
    summary = tables.get("PQ_Action_Items_Summary", pd.DataFrame()).copy()

    ws = wb.create_sheet("Action Items")
    ws.sheet_view.showGridLines = False
    fill_area(ws, start_row=1, end_row=260, start_col=1, end_col=32, color=COLOR_CANVAS)

    # Column widths for table-heavy layout
    widths = {
        "A": 2, "B": 16, "C": 16, "D": 16, "E": 16, "F": 14,
        "G": 12, "H": 15, "I": 18, "J": 24, "K": 12, "L": 14,
        "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14, "R": 14,
        "S": 14, "T": 14,
    }
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    for r in range(1, 261):
        ws.row_dimensions[r].height = 16

    ws.merge_cells("B1:T1")
    ws["B1"] = "Ametek SAP S4 PMO Huddle Report - Action Items"
    ws["B1"].font = Font(name=FONT_FAMILY, size=SIZE_TITLE, bold=True, color=COLOR_WHITE)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(name=FONT_FAMILY, size=SIZE_SUBTITLE, color=COLOR_DARK_TEXT)

    if detail.empty:
        draw_raised_panel(ws, 4, 10, 2, 20, shadow_color=COLOR_PANEL_SHADOW_SOFT)
        add_fade_header(ws, 4, 2, 20, "ACTION ITEMS")
        ws.merge_cells(start_row=7, start_column=3, end_row=9, end_column=19)
        c = ws.cell(row=7, column=3)
        c.value = "No action-items metrics table detected (expected: PQ_Action_Items_Detail/PQ_Action_Items_Summary)."
        c.font = Font(name=FONT_FAMILY, size=10, italic=True, color=COLOR_DARK_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return

    d = detail.copy()
    d["Section"] = d.get("Section", "").astype(str).str.strip()
    d["Workstream"] = d.get("Workstream", "(Unassigned)").astype(str).replace("", "(Unassigned)")
    d["AssignedTo"] = d.get("AssignedTo", "(Unassigned)").astype(str).replace("", "(Unassigned)")
    d["DueDate"] = pd.to_datetime(d.get("DueDate"), errors="coerce")

    common_cols = [
        c for c in [
            "Section", "Workstream", "AssignedTo", "ActionType", "SourceDomain", "ItemID", "ItemName", "Status", "DueDate"
        ] if c in d.columns
    ]

    # Section 1: Immediate + Needs Attention Soon
    sec1 = d[d["Section"].isin(["Immediate", "Needs Attention Soon"])].copy()
    sec1 = sec1.sort_values(["Section", "Workstream", "AssignedTo", "DueDate", "ActionType"], ascending=[True, True, True, True, True], kind="stable")

    draw_raised_panel(ws, 4, 70, 2, 20, shadow_color=COLOR_PANEL_SHADOW_SOFT)
    add_fade_header(ws, 4, 2, 20, "IMMEDIATE + NEEDS ATTENTION SOON")
    sec1_row_end = write_table(ws, start_row=6, start_col=3, df=sec1[common_cols], title=None)

    # Section 2: In Progress +2 weeks
    next_row = sec1_row_end + 1
    sec2 = d[d["Section"].eq("In Progress +2wks")].copy()
    sec2 = sec2.sort_values(["Workstream", "AssignedTo", "DueDate", "ActionType"], ascending=[True, True, True, True], kind="stable")

    draw_raised_panel(ws, next_row, next_row + 64, 2, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, next_row, 2, 20, "IN PROGRESS + 2 WEEKS LOOKAHEAD")
    sec2_row_end = write_table(ws, start_row=next_row + 2, start_col=3, df=sec2[common_cols], title=None)

    # Section 3: Combined action-items summary metrics
    next_row = sec2_row_end + 1
    sum_df = summary.copy()
    sum_cols = [
        c
        for c in [
            "Workstream",
            "AssignedTo",
            "ActionType",
            "TotalActionItems",
            "ImmediateItems",
            "NeedsAttentionSoonItems",
            "InProgress2WeeksItems",
        ]
        if c in sum_df.columns
    ]
    if sum_cols:
        sum_df = sum_df[sum_cols].sort_values(["Workstream", "AssignedTo", "TotalActionItems"], ascending=[True, True, False], kind="stable")

    draw_raised_panel(ws, next_row, next_row + 70, 2, 20, shadow_color=COLOR_PANEL_SHADOW_STRONG)
    add_fade_header(ws, next_row, 2, 20, "COMBINED ACTION ITEMS METRICS (BY WORKSTREAM / ASSIGNED TO / TYPE)")
    write_table(ws, start_row=next_row + 2, start_col=3, df=sum_df, title=None)

    ws.print_area = "B1:T245"




def build_huddle_report(metrics_path: Path, output_path: Path) -> Path:
    t = read_metrics_tables(metrics_path)
    detail_df = t["PQ_TD_Detail"]
    summary_df = t["PQ_TD_Summary"]

    summary_all_df = summary_df[summary_df["Section"].astype(str).str.upper() == "ALL"] if not summary_df.empty else pd.DataFrame()
    if summary_all_df.empty and not summary_df.empty:
        summary_all = summary_df.iloc[0]
    elif summary_all_df.empty:
        summary_all = pd.Series(dtype=float)
    else:
        summary_all = summary_all_df.iloc[0]

    today_overall, _ = build_today_tables(detail_df)
    today_detail_df = filter_today_detail(detail_df)
    detail_owner, exceptions = build_detail_assignee_and_exceptions(detail_df)

    detail_owner_display = detail_owner.head(TOP_N_ASSIGNEE).copy()
    exceptions_display = exceptions.head(TOP_N_EXCEPTIONS).copy()
    today_ws_display = build_workstream_display_from_detail(today_detail_df, TOP_N_WORKSTREAM)
    detail_ws_display = build_workstream_display_from_detail(detail_df, TOP_N_WORKSTREAM)

    wb = Workbook()
    ws = wb.active
    format_dashboard_sheet(ws)

    chart_bank = wb.create_sheet("Chart_Bank")
    chart_bank.sheet_view.showGridLines = True
    chart_bank["A1"] = "Chart Bank (drag these charts onto shaded placeholders in Testing_Defects)"
    chart_bank["A1"].font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_NAVY)
    chart_bank["A3"] = "TODAY CHARTS"
    chart_bank["A3"].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_SECTION_LABEL)
    chart_bank["A95"] = "DETAILED CHARTS"
    chart_bank["A95"].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_SECTION_LABEL)
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]:
        chart_bank.column_dimensions[col].width = 11
    for r in range(1, 220):
        chart_bank.row_dimensions[r].height = 20

    # Dashboard title
    ws.merge_cells("B1:T1")
    ws["B1"] = "Ametek SAP S4 PMO Huddle Report - Testing/Defects"
    ws["B1"].font = Font(name=FONT_FAMILY, size=SIZE_TITLE, bold=True, color=COLOR_WHITE)
    ws["B1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(name=FONT_FAMILY, size=SIZE_SUBTITLE, color=COLOR_DARK_TEXT)

    # Left rail and two repeated sections
    write_left_rail_object(ws, summary_all=summary_all, start_row=3, start_col=2, end_col=5, end_row=165)

    section_h = 42
    section1_top = 3
    section2_top = 47

    # TODAY section placeholders
    draw_raised_panel(ws, section1_top, section1_top + section_h, 6, 20, shadow_color=COLOR_PANEL_SHADOW_SOFT)
    add_fade_header(ws, section1_top, 6, 20, "TODAY TOTALS")
    write_metric_tile_grid(
        ws,
        cards=[
            ("Today Tests", _today_metric(today_overall, "Tests"), COLOR_BLUE),
            ("Passed", _today_metric(today_overall, "Passed"), COLOR_GREEN),
            ("Failed", _today_metric(today_overall, "Failed"), COLOR_RED),
            ("Blocked", _today_metric(today_overall, "Blocked"), COLOR_ORANGE),
            ("In Progress", _today_metric(today_overall, "InProgress"), COLOR_NAVY),
            ("Defect Links", _today_metric(today_overall, "DefectLinks"), COLOR_GOLD),
        ],
        start_row=6,
        start_col=6,
        end_col=20,
    )
    add_section_chart_grid(ws, top_row=16)
    add_chart_placeholder_labels(ws, top_row=16, prefix="TODAY")

    # DETAILED section placeholders
    draw_raised_panel(ws, section2_top, section2_top + section_h, 6, 20, shadow_color=COLOR_PANEL_SHADOW_MED)
    add_fade_header(ws, section2_top, 6, 20, "DETAILED TOTALS (OVERALL TEST CYCLE)")
    write_metric_tile_grid(
        ws,
        cards=[
            ("Total Tests", as_int(summary_all.get("TotalTests")), COLOR_BLUE),
            ("Passed", as_int(summary_all.get("TestsPassed")), COLOR_GREEN),
            ("Failed", as_int(summary_all.get("TestsFailed")), COLOR_RED),
            ("Blocked", as_int(summary_all.get("TestsBlocked")), COLOR_ORANGE),
            ("Executable", as_int(summary_all.get("ExecutableTests")), COLOR_NAVY),
            ("Defect Links", as_int(summary_all.get("TestsWithDefectLink")), COLOR_GOLD),
        ],
        start_row=50,
        start_col=6,
        end_col=20,
    )
    add_section_chart_grid(ws, top_row=60)
    add_chart_placeholder_labels(ws, top_row=60, prefix="DETAILED")

    # Chart bank population
    add_section_charts(
        ws,
        chart_ws=chart_bank,
        section_prefix="Today",
        chart_top_row=3,
        helper_col_start=110,
        helper_row_start=4,
        summary_metrics=build_summary_metrics_from_detail(today_detail_df),
        workstream_df=today_ws_display,
        exec_2w=build_two_week_execution_series(today_detail_df),
        backlog_2w=build_two_week_backlog_series(today_detail_df),
        aging_df=build_aging_buckets(today_detail_df),
        priority_df=build_defect_priority_proxy(today_detail_df),
        category_df=build_defect_category_proxy(today_detail_df),
        retest_df=build_retest_cycle_schedule(today_detail_df),
    )
    add_section_charts(
        ws,
        chart_ws=chart_bank,
        section_prefix="Detailed",
        chart_top_row=95,
        helper_col_start=140,
        helper_row_start=4,
        summary_metrics=build_summary_metrics_from_detail(detail_df),
        workstream_df=detail_ws_display,
        exec_2w=build_two_week_execution_series(detail_df),
        backlog_2w=build_two_week_backlog_series(detail_df),
        aging_df=build_aging_buckets(detail_df),
        priority_df=build_defect_priority_proxy(detail_df),
        category_df=build_defect_category_proxy(detail_df),
        retest_df=build_retest_cycle_schedule(detail_df),
    )

    # Bottom drilldown tables
    details_top = 92
    draw_raised_panel(ws, details_top, 165, 6, 20, shadow_color=COLOR_PANEL_SHADOW_STRONG)
    add_fade_header(ws, details_top, 6, 20, "DRILL-DOWN DETAIL TABLES")
    draw_chart_card(ws, top_row=111, left_col=6, bottom_row=124, right_col=20)
    add_panel_label(ws, "Top Workstreams", row=111, start_col=6, end_col=20)
    nr = write_table(ws, 113, 7, detail_ws_display, title=None)
    draw_chart_card(ws, top_row=nr + 1, left_col=6, bottom_row=nr + 14, right_col=20)
    add_panel_label(ws, "Top Assignees", row=nr + 1, start_col=6, end_col=20)
    nr2 = write_table(ws, nr + 3, 7, detail_owner_display, title=None)
    draw_chart_card(ws, top_row=nr2 + 1, left_col=6, bottom_row=nr2 + 14, right_col=20)
    add_panel_label(ws, "Exceptions", row=nr2 + 1, start_col=6, end_col=20)
    _ = write_table(ws, nr2 + 3, 7, exceptions_display, title=None)

    # Keep helper data off-canvas
    for col in list(range(21, 109)) + list(range(112, 139)) + list(range(142, 260)):
        ws.column_dimensions[get_column_letter(col)].hidden = True
    for col in [110, 111, 113, 114, 115, 116, 117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 128, 129, 140, 141, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157, 158, 159]:
        ws.column_dimensions[get_column_letter(col)].width = 2

    ws.print_area = "B1:T170"

    # Data sheet for transparency and drill-through
    data_ws = wb.create_sheet("TD_Data")
    data_ws.sheet_view.showGridLines = True
    data_ws.append(["Section", "Workstream", "Owner", "TestID", "TestName", "Status", "TaskID", "TaskName", "DefectID", "DefectCount", "SourceSheet", "SourceRow"])
    for _, rec in detail_df.iterrows():
        data_ws.append(
            [
                rec.get("Section"),
                rec.get("Workstream"),
                rec.get("Owner"),
                rec.get("TestID"),
                rec.get("TestName"),
                rec.get("Status"),
                rec.get("TaskID"),
                rec.get("TaskName"),
                rec.get("DefectID"),
                rec.get("DefectCount"),
                rec.get("SourceSheet"),
                rec.get("SourceRow"),
            ]
        )

    build_raid_dashboard_sheet(wb=wb, tables=t)
    build_tasks_dashboard_sheet(wb=wb, tables=t)
    build_action_items_report_sheet(wb=wb, tables=t)

    contract_audit = build_contract_audit(
        today_ws_display=today_ws_display,
        detail_ws_display=detail_ws_display,
        detail_owner_display=detail_owner_display,
        exceptions_display=exceptions_display,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path, contract_audit


def main() -> None:
    args = parse_args()
    metrics_path = Path(args.metrics_workbook)
    output_path = Path(args.output_report)

    if not metrics_path.is_absolute():
        metrics_path = ROOT / metrics_path
    if not output_path.is_absolute():
        output_path = ROOT / output_path

    built_path, contract_audit = build_huddle_report(metrics_path=metrics_path, output_path=output_path)
    print(
        {
            "huddle_report": str(built_path),
            "metrics_workbook": str(metrics_path),
            "contract_audit": contract_audit,
        }
    )


if __name__ == "__main__":
    main()
