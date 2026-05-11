from __future__ import annotations

import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


DEFAULT_METRICS = Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek PMO Metrics.xlsx")
DEFAULT_HUDDLE = Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek SAP S4 PMO Huddle Report.xlsx")

ACTION_ITEMS_EXPECTED_TABLES = ["PQ_Action_Items_Detail", "PQ_Action_Items_Summary"]
ACTION_ITEMS_EXPECTED_SUMMARY_COLUMNS = [
    "Workstream",
    "AssignedTo",
    "ActionType",
    "TotalActionItems",
    "ImmediateItems",
    "NeedsAttentionSoonItems",
    "InProgress2WeeksItems",
]
ACTION_ITEMS_EXPECTED_SECTIONS = [
    "IMMEDIATE + NEEDS ATTENTION SOON",
    "IN PROGRESS + 2 WEEKS LOOKAHEAD",
    "COMBINED ACTION ITEMS METRICS (BY WORKSTREAM / ASSIGNED TO / TYPE)",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate Tasks/RAID/Testing/Action Items dashboards against metrics workbook.")
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS, help="Path to Ametek PMO Metrics workbook")
    parser.add_argument("--huddle", type=Path, default=DEFAULT_HUDDLE, help="Path to Ametek PMO Huddle workbook")
    return parser.parse_args()


def _clean_text(v: Any) -> str:
    return str(v or "").strip()


def _to_num(v: Any) -> float:
    if v is None:
        return 0.0
    s = str(v).strip()
    if s == "":
        return 0.0
    s = s.replace(",", "")
    if s.endswith("%"):
        s = s[:-1].strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def _to_int(v: Any) -> int:
    return int(round(_to_num(v)))


def _normalize_bool(v: Any) -> bool:
    return str(v).strip().lower() in {"true", "1", "yes", "y"}


def _find_label_value(ws: openpyxl.worksheet.worksheet.Worksheet, labels: list[str], value_row_offset: int = 1, max_row: int = 80, max_col: int = 40) -> dict[str, Any]:
    found: dict[str, Any] = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True), start=1):
        for c_idx, raw in enumerate(row, start=1):
            cell_text = _clean_text(raw)
            if cell_text in labels and cell_text not in found:
                found[cell_text] = ws.cell(r_idx + value_row_offset, c_idx).value
    return found


def _find_section_headers(ws: openpyxl.worksheet.worksheet.Worksheet, targets: list[str], max_scan_rows: int = 4000, max_scan_cols: int = 30) -> dict[str, tuple[int, int]]:
    found: dict[str, tuple[int, int]] = {}
    max_row = min(max_scan_rows, max(1, int(ws.max_row or 1)))
    max_col = min(max_scan_cols, max(1, int(ws.max_column or 1)))

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True), start=1):
        for c_idx, raw in enumerate(row, start=1):
            val = _clean_text(raw).upper()
            if not val:
                continue
            for target in targets:
                if target in val and target not in found:
                    found[target] = (r_idx, c_idx)
                    if len(found) == len(targets):
                        return found
    return found


def validate_pmo_core(metrics_path: Path, huddle_path: Path) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []

    def add_check(name: str, actual: Any, expected: Any, tol: float | None = None) -> None:
        if tol is not None:
            passed = abs(_to_num(actual) - _to_num(expected)) <= tol
        else:
            passed = _to_int(actual) == _to_int(expected)
        checks.append({"name": name, "actual": actual, "expected": expected, "passed": passed})

    wbm = openpyxl.load_workbook(metrics_path, data_only=True, read_only=True)

    # RAID canonical from detail
    ws_rd = wbm["PQ_RAID_Detail"]
    rd_headers = [ws_rd.cell(1, c).value for c in range(1, ws_rd.max_column + 1)]
    rd_idx = {h: i + 1 for i, h in enumerate(rd_headers) if h}

    today = datetime.now()
    next7 = today + timedelta(days=7)

    raid_canon = {
        "Total": 0,
        "Open": 0,
        "Closed": 0,
        "Overdue": 0,
        "HighPriority": 0,
        "DueNext7": 0,
        "NeedsClarification": 0,
        "ImmediateAttention": 0,
    }

    for r in range(2, ws_rd.max_row + 1):
        raid_canon["Total"] += 1
        is_open = _normalize_bool(ws_rd.cell(r, rd_idx["IsOpen"]).value)
        is_overdue = _normalize_bool(ws_rd.cell(r, rd_idx["IsOverdue"]).value)
        is_high = _normalize_bool(ws_rd.cell(r, rd_idx["IsHighPriority"]).value)

        if is_open:
            raid_canon["Open"] += 1
            if is_overdue:
                raid_canon["Overdue"] += 1
            if is_high:
                raid_canon["HighPriority"] += 1

            dd_val = ws_rd.cell(r, rd_idx["DueDate"]).value
            if isinstance(dd_val, datetime) and today <= dd_val <= next7:
                raid_canon["DueNext7"] += 1

            owner = _clean_text(ws_rd.cell(r, rd_idx["Owner"]).value).lower()
            desc = _clean_text(ws_rd.cell(r, rd_idx["Description"]).value).lower()
            miti = _clean_text(ws_rd.cell(r, rd_idx["Mitigation"]).value).lower()
            comments = _clean_text(ws_rd.cell(r, rd_idx["Comments"]).value).lower()
            notes = _clean_text(ws_rd.cell(r, rd_idx["Notes"]).value).lower()
            stat = _clean_text(ws_rd.cell(r, rd_idx["Status"]).value).lower()

            needs_clar = (
                (owner in ["", "(unassigned)", "unassigned", "unknown", "tbd"])
                or desc == ""
                or miti == ""
                or comments == ""
                or notes == ""
                or any(x in stat for x in ["clarif", "need info", "needs info", "tbd", "unknown"])
            )
            if needs_clar:
                raid_canon["NeedsClarification"] += 1

            if is_overdue or is_high:
                raid_canon["ImmediateAttention"] += 1
        else:
            raid_canon["Closed"] += 1

    # RAID summary ALL
    ws_rs = wbm["PQ_RAID_Summary"]
    rs_headers = [ws_rs.cell(1, c).value for c in range(1, ws_rs.max_column + 1)]
    rs_idx = {h: i + 1 for i, h in enumerate(rs_headers) if h}
    all_row_raid = next((r for r in range(2, ws_rs.max_row + 1) if _clean_text(ws_rs.cell(r, 1).value) == "ALL"), None)
    if all_row_raid:
        add_check("RAID_Sum_Total", ws_rs.cell(all_row_raid, rs_idx["TotalItems"]).value, raid_canon["Total"])
        add_check("RAID_Sum_Open", ws_rs.cell(all_row_raid, rs_idx["OpenItems"]).value, raid_canon["Open"])
        add_check("RAID_Sum_Closed", ws_rs.cell(all_row_raid, rs_idx["ClosedItems"]).value, raid_canon["Closed"])
        add_check("RAID_Sum_Overdue", ws_rs.cell(all_row_raid, rs_idx["OverdueItems"]).value, raid_canon["Overdue"])
        add_check("RAID_Sum_HighPri", ws_rs.cell(all_row_raid, rs_idx["HighPriorityItems"]).value, raid_canon["HighPriority"])

    # Testing canonical from summary ALL
    ws_ts = wbm["PQ_TD_Summary"]
    ts_headers = [ws_ts.cell(1, c).value for c in range(1, ws_ts.max_column + 1)]
    ts_idx = {h: i + 1 for i, h in enumerate(ts_headers) if h}
    all_row_test = next((r for r in range(2, ws_ts.max_row + 1) if _clean_text(ws_ts.cell(r, 1).value) == "ALL"), None)
    td_canon: dict[str, Any] = {}
    if all_row_test:
        for h in ["TotalTests", "ExecutableTests", "TestsPassed", "TestsFailed", "TestsBlocked", "PassRatePct", "TestsWithDefectLink"]:
            td_canon[h] = ws_ts.cell(all_row_test, ts_idx[h]).value

    # Tasks canonical from summary ALL
    ws_task = wbm["PQ_Task_Summary"]
    task_headers = [ws_task.cell(1, c).value for c in range(1, ws_task.max_column + 1)]
    task_idx = {h: i + 1 for i, h in enumerate(task_headers) if h}
    all_row_task = next((r for r in range(2, ws_task.max_row + 1) if _clean_text(ws_task.cell(r, 1).value) == "ALL"), None)
    task_canon: dict[str, Any] = {}
    if all_row_task:
        for h in ["TotalTasks", "OpenTasks", "InProgressTasks", "OverdueTasks", "DueNext14", "ImmediateAttention", "PotentialRiskTasks"]:
            if h in task_idx:
                task_canon[h] = ws_task.cell(all_row_task, task_idx[h]).value

    wbm.close()

    # ----- Load huddle -----
    wbh = openpyxl.load_workbook(huddle_path, data_only=True, read_only=True)

    # RAID dashboard labels + values
    ws_rdsh = wbh["RAID Dashboard"]
    raid_labels = ["Total Items", "Open", "Closed", "Overdue", "Due Next 7", "High Priority", "Needs Clarification", "Immediate Attention"]
    raid_found = _find_label_value(ws_rdsh, raid_labels, value_row_offset=2, max_row=40, max_col=35)

    for lbl in raid_labels:
        checks.append({"name": f"RAID_Dash_Label_{lbl}", "actual": 1 if lbl in raid_found else 0, "expected": 1, "passed": lbl in raid_found})

    raid_map = {
        "Total Items": "Total",
        "Open": "Open",
        "Closed": "Closed",
        "Overdue": "Overdue",
        "Due Next 7": "DueNext7",
        "High Priority": "HighPriority",
        "Needs Clarification": "NeedsClarification",
        "Immediate Attention": "ImmediateAttention",
    }
    for lbl, key in raid_map.items():
        if lbl in raid_found:
            add_check(f"RAID_Dash_Value_{lbl}", raid_found[lbl], raid_canon[key])

    # Testing dashboard labels/values
    ws_td = wbh["Testing_Defects"]
    td_map = {
        "Total Tests": "TotalTests",
        "Executable": "ExecutableTests",
        "Passed": "TestsPassed",
        "Failed": "TestsFailed",
        "Blocked": "TestsBlocked",
        "Pass %": "PassRatePct",
        "Defect Links": "TestsWithDefectLink",
    }
    td_found = _find_label_value(ws_td, list(td_map.keys()), value_row_offset=1, max_row=80, max_col=12)
    for lbl in td_map:
        checks.append({"name": f"TD_Dash_Label_{lbl}", "actual": 1 if lbl in td_found else 0, "expected": 1, "passed": lbl in td_found})

    for lbl, key in td_map.items():
        if lbl in td_found and key in td_canon:
            if lbl == "Pass %":
                add_check(f"TD_Dash_Value_{lbl}", td_found[lbl], td_canon[key], tol=0.1)
            else:
                add_check(f"TD_Dash_Value_{lbl}", td_found[lbl], td_canon[key])

    # Tasks dashboard labels/values
    ws_tasks = wbh["Tasks Dashboard"]
    task_map = {
        "Total Tasks": "TotalTasks",
        "Open": "OpenTasks",
        "In Progress": "InProgressTasks",
        "Overdue": "OverdueTasks",
        "Due Next 14": "DueNext14",
        "Immediate Attention": "ImmediateAttention",
        "Potential Risk": "PotentialRiskTasks",
    }
    task_found = _find_label_value(ws_tasks, list(task_map.keys()), value_row_offset=2, max_row=40, max_col=35)
    for lbl in task_map:
        checks.append({"name": f"TASK_Dash_Label_{lbl}", "actual": 1 if lbl in task_found else 0, "expected": 1, "passed": lbl in task_found})

    for lbl, key in task_map.items():
        if lbl in task_found and key in task_canon:
            add_check(f"TASK_Dash_Value_{lbl}", task_found[lbl], task_canon[key])

    # milestone section marker
    milestone_found = False
    for row in ws_tasks.iter_rows(min_row=1, max_row=500, min_col=1, max_col=40, values_only=True):
        for raw in row:
            if "TASKS BY MILESTONE" in _clean_text(raw).upper():
                milestone_found = True
                break
        if milestone_found:
            break
    checks.append({"name": "TASK_Dash_Has_Milestone_Section", "actual": 1 if milestone_found else 0, "expected": 1, "passed": milestone_found})

    wbh.close()

    passed = sum(1 for c in checks if c["passed"])
    failed = len(checks) - passed
    verdict = "PASS" if failed == 0 else "FAIL"

    return {
        "verdict": verdict,
        "checks_total": len(checks),
        "checks_passed": passed,
        "checks_failed": failed,
        "raid_canonical": raid_canon,
        "testing_canonical": td_canon,
        "tasks_canonical": task_canon,
        "checks": checks,
    }


def validate_action_items(metrics_path: Path, huddle_path: Path) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []

    def add_check(name: str, passed: bool, detail: Any = "") -> None:
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    add_check("metrics_file_exists", metrics_path.exists(), str(metrics_path))
    add_check("huddle_file_exists", huddle_path.exists(), str(huddle_path))

    if not metrics_path.exists() or not huddle_path.exists():
        return {"verdict": "FAIL", "checks": checks}

    wb_metrics = openpyxl.load_workbook(metrics_path, data_only=True, read_only=True)
    metrics_sheets = wb_metrics.sheetnames
    for t in ACTION_ITEMS_EXPECTED_TABLES:
        add_check(f"metrics_has_{t}", t in metrics_sheets)
    wb_metrics.close()

    detail = pd.read_excel(metrics_path, sheet_name="PQ_Action_Items_Detail")
    summary = pd.read_excel(metrics_path, sheet_name="PQ_Action_Items_Summary")

    add_check("detail_has_rows", len(detail) > 0, len(detail))
    add_check("summary_has_rows", len(summary) > 0, len(summary))

    detail_has_section = "Section" in detail.columns
    add_check("detail_has_section_column", detail_has_section)
    section_counts: dict[str, int] = {}
    if detail_has_section:
        section_counts = detail["Section"].fillna("").astype(str).value_counts(dropna=False).to_dict()
        for section in ["Immediate", "Needs Attention Soon", "In Progress +2wks"]:
            add_check(f"detail_has_{section}", section in section_counts, section_counts)

    missing_summary_cols = [c for c in ACTION_ITEMS_EXPECTED_SUMMARY_COLUMNS if c not in summary.columns]
    add_check("summary_columns_complete", len(missing_summary_cols) == 0, missing_summary_cols)

    wb_huddle = openpyxl.load_workbook(huddle_path, data_only=True, read_only=True)
    add_check("huddle_has_action_items_sheet", "Action Items" in wb_huddle.sheetnames)

    section_positions: dict[str, tuple[int, int]] = {}
    if "Action Items" in wb_huddle.sheetnames:
        ws = wb_huddle["Action Items"]
        section_positions = _find_section_headers(ws, ACTION_ITEMS_EXPECTED_SECTIONS)
        for section in ACTION_ITEMS_EXPECTED_SECTIONS:
            add_check(f"huddle_has_section_{section}", section in section_positions, section_positions.get(section))

    wb_huddle.close()

    passed = sum(1 for c in checks if c["passed"])
    failed = len(checks) - passed
    verdict = "PASS" if failed == 0 else "FAIL"

    return {
        "verdict": verdict,
        "checks_total": len(checks),
        "checks_passed": passed,
        "checks_failed": failed,
        "detail_rows": int(len(detail)),
        "summary_rows": int(len(summary)),
        "section_counts": section_counts,
        "section_positions": section_positions,
        "checks": checks,
    }


def main() -> None:
    args = parse_args()
    metrics_path = args.metrics
    huddle_path = args.huddle

    core = validate_pmo_core(metrics_path, huddle_path)
    action_items = validate_action_items(metrics_path, huddle_path)

    overall_verdict = "PASS" if core.get("verdict") == "PASS" and action_items.get("verdict") == "PASS" else "FAIL"

    payload = {
        "verdict": overall_verdict,
        "metrics": str(metrics_path),
        "huddle": str(huddle_path),
        "pmo_core": core,
        "action_items": action_items,
    }
    print(json.dumps(payload, indent=2, default=str))


if __name__ == "__main__":
    main()
