"""
Build metrics.xlsx with proper Excel Table structure:
  - One Summary + Details pair per domain
  - Formulas that reference Details tables
  - Efficient data structure (limits details to 500 rows for efficiency)
  - Ready for Power Query connections
  
Architecture:
  Resources_Summary → queries Resources_Details
  Resources_Details → consolidated from project plan resources
  [Domain]_Summary → queries [Domain]_Details (Risks, Issues, Actions)
  [Domain]_Details → full domain data
"""

import pandas as pd
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def get_source_paths():
    """Get correct paths to source files"""
    base = Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\SAP Implementation Project - Documents")
    return {
        'project_plan': base / "00 - PMO" / "02 - Project Planning" / "AMETEK SAP S4 Master Project Plan 3ITC May EXCEL.xlsx",
        'raid_log': base / "00 - PMO" / "04 - Project Tracking and Reporting" / "AMETEK SAP S4 RAID Log.xlsm",
    }

def load_project_plan():
    """Load project plan and extract key columns for efficiency"""
    sources = get_source_paths()
    try:
        df = pd.read_excel(sources['project_plan'], sheet_name='Project Plan', header=0)
        print(f"✓ Project Plan loaded: {len(df)} rows, {len(df.columns)} columns")
        
        # Extract key columns only (not all 100+ columns)
        key_cols = ['Task Name', 'Status', 'Resource Names', 'Start', 'Finish', 
                   'Due Date', '% Complete', 'Workstream', 'Duration', 'Milestone']
        
        # Find which columns exist
        available = [c for c in key_cols if c in df.columns]
        df_slim = df[available].copy() if available else df.iloc[:, :10]
        
        print(f"  Extracted {len(df_slim.columns)} key columns")
        return df_slim
    except Exception as e:
        print(f"✗ Project Plan error: {e}")
        return None

def load_raid_log():
    """Load RAID log data - all 4 sheets"""
    sources = get_source_paths()
    data = {}
    
    try:
        df_risks = pd.read_excel(sources['raid_log'], sheet_name='Risks')
        data['risks'] = df_risks
        print(f"✓ Risks: {len(df_risks)} rows")
    except Exception as e:
        print(f"✗ Risks: {e}")
    
    try:
        df_issues = pd.read_excel(sources['raid_log'], sheet_name='Issues')
        data['issues'] = df_issues
        print(f"✓ Issues: {len(df_issues)} rows")
    except Exception as e:
        print(f"✗ Issues: {e}")
    
    try:
        df_decisions = pd.read_excel(sources['raid_log'], sheet_name='Decisions')
        data['decisions'] = df_decisions
        print(f"✓ Decisions: {len(df_decisions)} rows")
    except Exception as e:
        print(f"✗ Decisions: {e}")
    
    try:
        df_actions = pd.read_excel(sources['raid_log'], sheet_name='Actions')
        data['actions'] = df_actions
        print(f"✓ Actions: {len(df_actions)} rows")
    except Exception as e:
        print(f"✗ Actions: {e}")
    
    return data

def create_resources_tables(df_plan):
    """Extract Resources Summary and Details from project plan"""
    if df_plan is None:
        return None, None
    
    # Get resource column (try common names)
    resource_col = None
    for col in ['Resource Names', 'Resource', 'Assigned To', 'Owner']:
        if col in df_plan.columns:
            resource_col = col
            break
    
    if not resource_col:
        print("  ✗ No resource column found")
        print(f"  Available columns: {list(df_plan.columns[:20])}")
        return None, None
    
    # Build details table (limit to 200 rows for efficiency)
    details = []
    for idx, row in df_plan.iterrows():
        if len(details) >= 200:  # Limit to 200 rows
            break
        resources_str = str(row.get(resource_col, ''))
        if pd.notna(resources_str) and resources_str.strip():
            # Split semicolon-separated resources
            for resource in resources_str.split(';'):
                resource = resource.strip()
                if resource and resource.lower() not in ['unassigned', 'tbd', '']:
                    if len(details) < 200:  # Check again before adding
                        details.append({
                            'Person': resource,
                            'Task': row.get('Task', row.get('Task Name', '')),
                            'Status': row.get('Status', ''),
                            'Due Date': row.get('Due Date', row.get('Finish', '')),
                            'Completion %': row.get('% Complete', 0),
                        })
    
    df_details = pd.DataFrame(details) if details else pd.DataFrame()
    
    # Create summary table
    if len(df_details) > 0:
        df_summary = df_details.groupby('Person').agg({
            'Task': 'count',
            'Status': lambda x: (x.str.lower() == 'closed').sum() if hasattr(x, 'str') else 0,
            'Completion %': 'mean',
        }).reset_index()
        df_summary.columns = ['Person', 'Total Tasks', 'Completed', 'Avg Completion %']
        df_summary['Open'] = df_summary['Total Tasks'] - df_summary['Completed']
    else:
        df_summary = pd.DataFrame({'Person': [], 'Total Tasks': [], 'Completed': [], 'Avg Completion %': [], 'Open': []})
    
    print(f"✓ Resources: {len(df_summary)} people, {len(df_details)} task assignments")
    return df_summary, df_details

def create_project_plan_tables(df_plan):
    """Extract Project Plan Summary and Details"""
    if df_plan is None:
        return None, None
    
    # Build details table (limit to 300 rows for efficiency)
    details = df_plan[['Task Name', 'Status', 'Start', 'Finish', '% Complete', 'Duration', 'Milestone']].head(300).copy()
    
    # Create summary table
    summary = {
        'Total Tasks': len(df_plan),
        'Closed': (df_plan['Status'].str.lower() == 'closed').sum() if 'Status' in df_plan.columns else 0,
        'Active': (df_plan['Status'].str.lower() != 'closed').sum() if 'Status' in df_plan.columns else len(df_plan),
        'Completion %': df_plan.get('% Complete', 0).mean() if '% Complete' in df_plan.columns else 0,
        'Milestones': (df_plan['Milestone'] == True).sum() if 'Milestone' in df_plan.columns else 0,
    }
    df_summary = pd.DataFrame([summary])
    
    print(f"✓ Project Plan: {len(df_summary)} summary, {len(details)} details")
    return df_summary, details


def create_testing_defects_tables():
    """Create Testing and Defects Summary tables (stubs for now)"""
    # Testing summary (will be populated from source)
    testing_summary = {
        'Total Tests': 0,
        'Passed': 0,
        'Failed': 0,
        'Blocked': 0,
        'Not Run': 0,
    }
    df_testing_summary = pd.DataFrame([testing_summary])
    df_testing_details = pd.DataFrame({'Test ID': [], 'Status': [], 'Result': [], 'Owner': []})
    
    # Defects summary
    defects_summary = {
        'Total Defects': 0,
        'Open': 0,
        'Closed': 0,
        'Critical': 0,
    }
    df_defects_summary = pd.DataFrame([defects_summary])
    df_defects_details = pd.DataFrame({'Defect ID': [], 'Status': [], 'Severity': [], 'Owner': []})
    
    print(f"✓ Testing: 1 summary, 0 details (stub - add data source)")
    print(f"✓ Defects: 1 summary, 0 details (stub - add data source)")
    
    return df_testing_summary, df_testing_details, df_defects_summary, df_defects_details


def create_summary_tables(raid_data):
    """Create summary tables for each domain"""
    summaries = {}
    
    for domain in ['risks', 'issues', 'decisions', 'actions']:
        if domain not in raid_data:
            continue
        
        df = raid_data[domain]
        # Count by status
        status_col = None
        for col in df.columns:
            if col.lower() in ['status', 'state']:
                status_col = col
                break
        
        if status_col:
            summary = {
                'Total': len(df),
                'Open': (df[status_col].str.lower() == 'open').sum() if hasattr(df[status_col], 'str') else 0,
                'Closed': (df[status_col].str.lower() == 'closed').sum() if hasattr(df[status_col], 'str') else 0,
            }
        else:
            summary = {'Total': len(df), 'Open': len(df), 'Closed': 0}
        
        summaries[domain] = pd.DataFrame([summary])
    
    return summaries

def write_workbook(output_path):
    """Write metrics workbook with formulas built in"""
    print(f"Writing to {output_path}...")
    
    # Load data
    df_plan = load_project_plan()
    raid_data = load_raid_log()
    
    # Build details dataframes
    _, res_details = create_resources_tables(df_plan)
    _, plan_details = create_project_plan_tables(df_plan)
    _, test_details, _, defect_details = create_testing_defects_tables()

    details_frames = {
        'Resources_Details': res_details if res_details is not None else pd.DataFrame({'Person': [], 'Task': [], 'Status': [], 'Due Date': [], 'Completion %': []}),
        'ProjectPlan_Details': plan_details if plan_details is not None else pd.DataFrame({'Task Name': [], 'Status': [], 'Start': [], 'Finish': [], '% Complete': [], 'Duration': [], 'Milestone': []}),
        'Testing_Details': test_details,
        'Defects_Details': defect_details,
        'Risks_Details': raid_data['risks'].head(500) if 'risks' in raid_data else pd.DataFrame(),
        'Issues_Details': raid_data['issues'].head(500) if 'issues' in raid_data else pd.DataFrame(),
        'Decisions_Details': raid_data['decisions'].head(500) if 'decisions' in raid_data else pd.DataFrame(),
        'Actions_Details': raid_data['actions'].head(500) if 'actions' in raid_data else pd.DataFrame(),
    }

    # Build summary metric rows (Metric/Value only)
    summary_rows = {
        'Resources_Summary': ['Total Resources', 'Total Assignments', 'Avg Completion %', 'On Track (>75%)'],
        'ProjectPlan_Summary': ['Total Tasks', 'Complete', 'Future/On Schedule/Late', 'Avg Completion %'],
        'Testing_Summary': ['Total Tests', 'Passed', 'Failed', 'Blocked'],
        'Defects_Summary': ['Total Defects', 'Open', 'Closed', 'Critical'],
        'Risks_Summary': ['Total Risks', 'Open', 'Closed'],
        'Issues_Summary': ['Total Issues', 'Open + Escalated', 'Closed'],
        'Decisions_Summary': ['Total Decisions', 'Pending', 'Approved'],
        'Actions_Summary': ['Total Actions', 'Open + In Progress', 'Closed'],
    }

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, metrics in summary_rows.items():
            pd.DataFrame({'Metric': metrics, 'Value': [None] * len(metrics)}).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        for sheet_name, df_details in details_frames.items():
            df_details.to_excel(writer, sheet_name=sheet_name, index=False)

    # Add formulas and tables
    wb = load_workbook(output_path)

    # Summary formulas
    wb['Resources_Summary']['B2'] = '=COUNTA(ResourcesDetails[Person])-COUNTBLANK(ResourcesDetails[Person])'
    wb['Resources_Summary']['B3'] = '=COUNTA(ResourcesDetails[Task])'
    wb['Resources_Summary']['B4'] = '=AVERAGE(ResourcesDetails[Completion %])'
    wb['Resources_Summary']['B5'] = '=COUNTIF(ResourcesDetails[Completion %],">75%")'

    wb['ProjectPlan_Summary']['B2'] = '=COUNTA(ProjectPlanDetails[Task Name])'
    wb['ProjectPlan_Summary']['B3'] = '=COUNTIF(ProjectPlanDetails[Status],"Complete")'
    wb['ProjectPlan_Summary']['B4'] = '=COUNTIFS(ProjectPlanDetails[Status],"<>Complete",ProjectPlanDetails[Task Name],"<>")'
    wb['ProjectPlan_Summary']['B5'] = '=AVERAGE(ProjectPlanDetails[% Complete])'

    wb['Testing_Summary']['B2'] = '=COUNTA(TestingDetails[Test ID])'
    wb['Testing_Summary']['B3'] = '=COUNTIF(TestingDetails[Status],"Passed")'
    wb['Testing_Summary']['B4'] = '=COUNTIF(TestingDetails[Status],"Failed")'
    wb['Testing_Summary']['B5'] = '=COUNTIF(TestingDetails[Status],"Blocked")'

    wb['Defects_Summary']['B2'] = '=COUNTA(DefectsDetails[Defect ID])'
    wb['Defects_Summary']['B3'] = '=COUNTIF(DefectsDetails[Status],"Open")'
    wb['Defects_Summary']['B4'] = '=COUNTIF(DefectsDetails[Status],"Closed")'
    wb['Defects_Summary']['B5'] = '=COUNTIF(DefectsDetails[Severity],"Critical")'

    wb['Risks_Summary']['B2'] = '=COUNTA(RisksDetails[Risk ID])'
    wb['Risks_Summary']['B3'] = '=COUNTIF(RisksDetails[Risk Status],"Open")'
    wb['Risks_Summary']['B4'] = '=COUNTIF(RisksDetails[Risk Status],"Closed")'
    # Note: blank rows are filtered in PQ_Risks via Text.StartsWith([Risk ID],"R-")

    wb['Issues_Summary']['B2'] = '=COUNTA(IssuesDetails[Issue ID])'
    wb['Issues_Summary']['B3'] = '=COUNTIF(IssuesDetails[Issue Status],"Open")+COUNTIF(IssuesDetails[Issue Status],"Escalated")'
    wb['Issues_Summary']['B4'] = '=COUNTIF(IssuesDetails[Issue Status],"Closed")'

    wb['Decisions_Summary']['B2'] = '=COUNTA(DecisionsDetails[Decision ID])'
    wb['Decisions_Summary']['B3'] = '=COUNTIF(DecisionsDetails[Decision Status],"Pending")'
    wb['Decisions_Summary']['B4'] = '=COUNTIF(DecisionsDetails[Decision Status],"Approved")'

    wb['Actions_Summary']['B2'] = '=COUNTA(ActionsDetails[Action Item ID])'
    wb['Actions_Summary']['B3'] = '=COUNTIF(ActionsDetails[Action Item Status],"Open")+COUNTIF(ActionsDetails[Action Item Status],"In Progress")'
    wb['Actions_Summary']['B4'] = '=COUNTIF(ActionsDetails[Action Item Status],"Closed")'

    # Create valid Excel Tables for all sheets
    table_specs = [
        ('Resources_Summary', 'ResourcesSummary', 'TableStyleMedium9'),
        ('Resources_Details', 'ResourcesDetails', 'TableStyleMedium2'),
        ('ProjectPlan_Summary', 'ProjectPlanSummary', 'TableStyleMedium9'),
        ('ProjectPlan_Details', 'ProjectPlanDetails', 'TableStyleMedium2'),
        ('Testing_Summary', 'TestingSummary', 'TableStyleMedium9'),
        ('Testing_Details', 'TestingDetails', 'TableStyleMedium2'),
        ('Defects_Summary', 'DefectsSummary', 'TableStyleMedium9'),
        ('Defects_Details', 'DefectsDetails', 'TableStyleMedium2'),
        ('Risks_Summary', 'RisksSummary', 'TableStyleMedium9'),
        ('Risks_Details', 'RisksDetails', 'TableStyleMedium2'),
        ('Issues_Summary', 'IssuesSummary', 'TableStyleMedium9'),
        ('Issues_Details', 'IssuesDetails', 'TableStyleMedium2'),
        ('Decisions_Summary', 'DecisionsSummary', 'TableStyleMedium9'),
        ('Decisions_Details', 'DecisionsDetails', 'TableStyleMedium2'),
        ('Actions_Summary', 'ActionsSummary', 'TableStyleMedium9'),
        ('Actions_Details', 'ActionsDetails', 'TableStyleMedium2'),
    ]

    for sheet_name, table_name, style_name in table_specs:
        ws = wb[sheet_name]

        if ws.max_row < 2:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=2, column=col, value=None)

        end_col = get_column_letter(ws.max_column)
        table_ref = f"A1:{end_col}{ws.max_row}"

        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    wb.save(output_path)
    
    print(f"✓ Workbook created with formulas: {output_path}")
    
    # List sheets
    print("\nWorksheet structure:")
    print("  ✓ Resources_Summary/Details")
    print("  ✓ ProjectPlan_Summary/Details")
    print("  ✓ Testing_Summary/Details")
    print("  ✓ Defects_Summary/Details")
    print("  ✓ Risks_Summary/Details")
    print("  ✓ Issues_Summary/Details")
    print("  ✓ Decisions_Summary/Details")
    print("  ✓ Actions_Summary/Details")

def main():
    output_dir = Path(r'C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report')
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / 'metrics.xlsx'
    
    # Remove old file
    if output_path.exists():
        output_path.unlink()
    
    write_workbook(output_path)
    
    print("\n" + "=" * 70)
    print("METRICS WORKBOOK CREATED WITH SUMMARY + DETAILS STRUCTURE")
    print("=" * 70)
    print("\nNext steps:")
    print("  1. Open metrics.xlsx in Excel")
    print("  2. Verify Summary formulas calculate from Details tables")
    print("  3. Add/refresh Power Query connections to source files")

if __name__ == '__main__':
    main()
