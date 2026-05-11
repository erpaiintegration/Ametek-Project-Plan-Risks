from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
import pandas as pd


DEFAULT_METRICS = Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek PMO Metrics.xlsx")
DEFAULT_HUDDLE = Path(r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek SAP S4 PMO Huddle Report.xlsx")

EXPECTED_TABLES = ["PQ_Action_Items_Detail", "PQ_Action_Items_Summary"]
EXPECTED_SUMMARY_COLUMNS = [
    "Workstream",
    "AssignedTo",
    "ActionType",
    "TotalActionItems",
    "ImmediateItems",
    "NeedsAttentionSoonItems",
    "InProgress2WeeksItems",
]
EXPECTED_SECTIONS = [
    "IMMEDIATE + NEEDS ATTENTION SOON",
    "IN PROGRESS + 2 WEEKS LOOKAHEAD",
    "COMBINED ACTION ITEMS METRICS (BY WORKSTREAM / ASSIGNED TO / TYPE)",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate Action Items metrics tables and huddle sheet layout.")
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS, help="Path to Ametek PMO Metrics workbook")
    parser.add_argument("--huddle", type=Path, default=DEFAULT_HUDDLE, help="Path to Ametek PMO Huddle workbook")
    return parser.parse_args()


def _normalize_cell(v: object) -> str:
    return str(v or "").strip()


def _find_section_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    targets: list[str],
    max_scan_rows: int = 4000,
    max_scan_cols: int = 30,
) -> dict[str, tuple[int, int]]:
    found: dict[str, tuple[int, int]] = {}
    max_row = min(max_scan_rows, max(1, int(ws.max_row or 1)))
    max_col = min(max_scan_cols, max(1, int(ws.max_column or 1)))

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
        start=1,
    ):
        for c_idx, raw in enumerate(row, start=1):
            val = _normalize_cell(raw).upper()
            if not val:
                continue
            for target in targets:
                if target in val and target not in found:
                    found[target] = (r_idx, c_idx)
                    if len(found) == len(targets):
                        return found
    return found


def main() -> None:
    args = parse_args()
    metrics_path = args.metrics
    huddle_path = args.huddle

    checks: list[dict[str, object]] = []

    def add_check(name: str, passed: bool, detail: object = "") -> None:
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    # --- Metrics workbook checks ---
    add_check("metrics_file_exists", metrics_path.exists(), str(metrics_path))
    add_check("huddle_file_exists", huddle_path.exists(), str(huddle_path))

    if not metrics_path.exists() or not huddle_path.exists():
        print(json.dumps({"verdict": "FAIL", "checks": checks}, indent=2))
        return

    wb_metrics = openpyxl.load_workbook(metrics_path, data_only=True, read_only=True)
    metrics_sheets = wb_metrics.sheetnames

    for t in EXPECTED_TABLES:
        add_check(f"metrics_has_{t}", t in metrics_sheets)

    wb_metrics.close()

    detail = pd.read_excel(metrics_path, sheet_name="PQ_Action_Items_Detail")
    summary = pd.read_excel(metrics_path, sheet_name="PQ_Action_Items_Summary")

    add_check("detail_has_rows", len(detail) > 0, len(detail))
    add_check("summary_has_rows", len(summary) > 0, len(summary))

    detail_has_section = "Section" in detail.columns
    add_check("detail_has_section_column", detail_has_section)

    if detail_has_section:
        section_counts = detail["Section"].fillna("").astype(str).value_counts(dropna=False).to_dict()
        for section in ["Immediate", "Needs Attention Soon", "In Progress +2wks"]:
            add_check(f"detail_has_{section}", section in section_counts, section_counts)
    else:
        section_counts = {}

    missing_summary_cols = [c for c in EXPECTED_SUMMARY_COLUMNS if c not in summary.columns]
    add_check("summary_columns_complete", len(missing_summary_cols) == 0, missing_summary_cols)

    # --- Huddle workbook checks ---
    wb_huddle = openpyxl.load_workbook(huddle_path, data_only=True, read_only=True)
    add_check("huddle_has_action_items_sheet", "Action Items" in wb_huddle.sheetnames)

    section_positions: dict[str, tuple[int, int]] = {}
    if "Action Items" in wb_huddle.sheetnames:
        ws = wb_huddle["Action Items"]
        section_positions = _find_section_headers(ws, EXPECTED_SECTIONS)
        for section in EXPECTED_SECTIONS:
            add_check(f"huddle_has_section_{section}", section in section_positions, section_positions.get(section))

    wb_huddle.close()

    passed = sum(1 for c in checks if c["passed"])
    failed = len(checks) - passed
    verdict = "PASS" if failed == 0 else "FAIL"

    output = {
        "verdict": verdict,
        "checks_total": len(checks),
        "checks_passed": passed,
        "checks_failed": failed,
        "detail_rows": int(len(detail)),
        "summary_rows": int(len(summary)),
        "section_counts": section_counts,
        "section_positions": section_positions,
        "checks": checks,
    }

    print(json.dumps(output, indent=2, default=str))


if __name__ == "__main__":
    main()
