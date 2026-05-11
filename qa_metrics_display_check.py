import openpyxl
from collections import defaultdict

metrics_path = r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek PMO Metrics.xlsx"
huddle_path = r"C:\Users\jsw73\OneDrive - AMETEK Inc\Ametek SAP S4 Daily Report\Ametek SAP S4 PMO Huddle Report.xlsx"

results = []
def check(name, ok, detail=""):
    results.append((name, bool(ok), detail))

def norm_num(v):
    if v is None:
        return 0
    s = str(v).strip()
    if s == "":
        return 0
    try:
        return int(float(s))
    except Exception:
        return 0

# Load metrics workbook
wbm = openpyxl.load_workbook(metrics_path, data_only=True, read_only=True)
expected_sheets = [
    'PQ_Task_Detail','PQ_Task_Summary','PQ_Task_Workstream','PQ_Task_Milestone',
    'PQ_RAID_Detail','PQ_RAID_Summary','PQ_RAID_Workstream',
    'PQ_TD_Detail','PQ_TD_Summary','PQ_TD_Workstream'
]
for s in expected_sheets:
    check(f"metrics_sheet:{s}", s in wbm.sheetnames)

# --- TASKS reconciliation ---
wsd = wbm['PQ_Task_Detail']
headers = [wsd.cell(1,c).value for c in range(1,wsd.max_column+1)]
idx = {h:i+1 for i,h in enumerate(headers) if h}
required_task_cols = ['TaskID','Workstream','Milestone','IsOpen','IsComplete','IsInProgress','IsOverdue','IsDueNext14','IsDueNext7','IsImmediateAttention','IsPotentialRisk']
for c in required_task_cols:
    check(f"tasks_detail_col:{c}", c in idx)

agg = defaultdict(int)
by_ws = defaultdict(lambda: defaultdict(int))
by_ms = defaultdict(lambda: defaultdict(int))
rows_detail = 0
for r in range(2, wsd.max_row+1):
    task_id = wsd.cell(r, idx['TaskID']).value if 'TaskID' in idx else None
    if task_id in (None, ""):
        continue
    rows_detail += 1
    ws = str(wsd.cell(r, idx['Workstream']).value or '(Unassigned)').strip()
    ms = str(wsd.cell(r, idx['Milestone']).value or '(Unassigned)').strip()
    vals = {
        'TotalTasks': 1,
        'OpenTasks': norm_num(wsd.cell(r, idx['IsOpen']).value),
        'CompleteTasks': norm_num(wsd.cell(r, idx['IsComplete']).value),
        'InProgressTasks': norm_num(wsd.cell(r, idx['IsInProgress']).value),
        'OverdueTasks': norm_num(wsd.cell(r, idx['IsOverdue']).value),
        'DueNext14': norm_num(wsd.cell(r, idx['IsDueNext14']).value),
        'DueNext7': norm_num(wsd.cell(r, idx['IsDueNext7']).value),
        'ImmediateAttention': norm_num(wsd.cell(r, idx['IsImmediateAttention']).value),
        'PotentialRiskTasks': norm_num(wsd.cell(r, idx['IsPotentialRisk']).value),
    }
    for k,v in vals.items():
        agg[k] += v
        by_ws[ws][k] += v
        by_ms[ms][k] += v

check("tasks_detail_rows_positive", rows_detail > 0, f"rows={rows_detail}")

# Compare with summary ALL
wss = wbm['PQ_Task_Summary']
sh = [wss.cell(1,c).value for c in range(1,wss.max_column+1)]
sidx = {h:i+1 for i,h in enumerate(sh) if h}
all_row = None
for r in range(2, wss.max_row+1):
    if str(wss.cell(r,1).value).strip() == 'ALL':
        all_row = r
        break
check("tasks_summary_all_exists", all_row is not None)

metrics = ['TotalTasks','OpenTasks','CompleteTasks','InProgressTasks','OverdueTasks','DueNext14','DueNext7','ImmediateAttention','PotentialRiskTasks']
if all_row:
    for m in metrics:
        expected = agg[m]
        actual = norm_num(wss.cell(all_row, sidx[m]).value) if m in sidx else None
        check(f"tasks_all_reconcile:{m}", actual == expected, f"actual={actual}, expected={expected}")

# Compare workstream summary rows
wsws = wbm['PQ_Task_Workstream']
wh = [wsws.cell(1,c).value for c in range(1,wsws.max_column+1)]
widx = {h:i+1 for i,h in enumerate(wh) if h}
status_col_ok = 'Status' in widx
check('tasks_workstream_has_status', status_col_ok)
# aggregate workstream table to ws totals
ws_agg = defaultdict(lambda: defaultdict(int))
for r in range(2, wsws.max_row+1):
    ws = str(wsws.cell(r, widx['Workstream']).value or '(Unassigned)').strip()
    if ws == 'ALL' or ws == '':
        continue
    for m in metrics:
        if m in widx:
            ws_agg[ws][m] += norm_num(wsws.cell(r, widx[m]).value)

for ws, mvals in by_ws.items():
    for m in metrics:
        check(f"tasks_workstream_reconcile:{ws}:{m}", ws_agg[ws][m] == mvals[m], f"actual={ws_agg[ws][m]}, expected={mvals[m]}")

# Compare milestone table
wsm = wbm['PQ_Task_Milestone']
mh = [wsm.cell(1,c).value for c in range(1,wsm.max_column+1)]
midx = {h:i+1 for i,h in enumerate(mh) if h}
for c in ['Milestone','TotalTasks','OpenTasks','PotentialRiskTasks']:
    check(f"tasks_milestone_col:{c}", c in midx)

ms_agg = defaultdict(lambda: defaultdict(int))
for r in range(2, wsm.max_row+1):
    ms = str(wsm.cell(r, midx['Milestone']).value or '(Unassigned)').strip()
    for m in metrics:
        if m in midx:
            ms_agg[ms][m] += norm_num(wsm.cell(r, midx[m]).value)

for ms, mvals in by_ms.items():
    for m in metrics:
        if m in midx:
            check(f"tasks_milestone_reconcile:{ms}:{m}", ms_agg[ms][m] == mvals[m], f"actual={ms_agg[ms][m]}, expected={mvals[m]}")

wbm.close()

# --- DISPLAY QA ---
wbh = openpyxl.load_workbook(huddle_path, data_only=True, read_only=True)
check('display_sheet:Tasks Dashboard', 'Tasks Dashboard' in wbh.sheetnames)
if 'Tasks Dashboard' in wbh.sheetnames:
    ws = wbh['Tasks Dashboard']
    kpis = {}
    for c in range(1, 35):
        lbl = ws.cell(9,c).value
        if lbl is not None and str(lbl).strip() != '':
            kpis[str(lbl).strip()] = ws.cell(11,c).value

    for required_lbl in ['Total Tasks','Open','In Progress','Overdue','Due Next 14','Immediate Attention','Potential Risk']:
        check(f"display_kpi_present:{required_lbl}", required_lbl in kpis)

    expected_map = {
        'Total Tasks': agg['TotalTasks'],
        'Open': agg['OpenTasks'],
        'In Progress': agg['InProgressTasks'],
        'Overdue': agg['OverdueTasks'],
        'Due Next 14': agg['DueNext14'],
        'Immediate Attention': agg['ImmediateAttention'],
        'Potential Risk': agg['PotentialRiskTasks'],
    }
    for lbl, exp in expected_map.items():
        act = norm_num(kpis.get(lbl))
        check(f"display_kpi_reconcile:{lbl}", act == exp, f"actual={act}, expected={exp}")

    has_milestone_section = False
    for r in range(1, 320):
        val = ws.cell(r,2).value
        if isinstance(val, str) and 'TASKS BY MILESTONE' in val.upper():
            has_milestone_section = True
            break
    check('display_section_tasks_by_milestone', has_milestone_section)

wbh.close()

# Print results
fails = [x for x in results if not x[1]]
passes = [x for x in results if x[1]]
print(f"QA_TOTAL={len(results)}")
print(f"QA_PASS={len(passes)}")
print(f"QA_FAIL={len(fails)}")
if fails:
    print("QA_FAILED_CHECKS_START")
    for n,ok,d in fails[:200]:
        print(n, "::", d)
    print("QA_FAILED_CHECKS_END")
else:
    print("QA_ALL_CHECKS_PASSED")

print("TASKS_SUMMARY_ALL", {k: agg[k] for k in ['TotalTasks','OpenTasks','InProgressTasks','OverdueTasks','DueNext14','ImmediateAttention','PotentialRiskTasks']})
